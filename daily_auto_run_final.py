"""
daily_auto_run_final.py  (v6.3.29)
新增：把「動態權重」同步輸出到每日 Excel（weight_* 欄位可追溯）。
"""
from __future__ import annotations
import datetime as _dt





import math

def _as_float(x):
    try:
        if x is None or (isinstance(x, float) and math.isnan(x)):
            return None
        s = str(x).strip()
        if s == "" or s.lower() in ("nan","none","n/a"):
            return None
        return float(s)
    except Exception:
        return None

def _emoji_light(v, t1, t2, reverse=False):
    if v is None:
        return "N/A"
    if reverse:
        if v >= t2: return "🔴"
        if v >= t1: return "🟡"
        return "🟢"
    else:
        if v >= t2: return "🟢"
        if v >= t1: return "🟡"
        return "🔴"

SMR_LOW_LABEL = "LOW"
SMR_LOW_THRESHOLD = 9.0

def ensure_lights_and_squeeze(df):
    """Compute lights & squeeze pressure robustly (independent of strategy_score)."""
    if df is None or len(df)==0:
        return df
    import pandas as pd
    out = df.copy()

    # Turnover light
    tcol = None
    for c in ["周轉率(%)","turnover_rate(%)","turnover_rate"]:
        if c in out.columns:
            tcol = c; break
    if "周轉率燈號" not in out.columns:
        out["周轉率燈號"] = ""
    if tcol:
        tr = pd.to_numeric(out[tcol], errors="coerce")
        out["周轉率燈號"] = tr.map(lambda x: _emoji_light(_as_float(x), 0.2, 0.5, reverse=False))

    # Trade value light
    if "成交值燈號" not in out.columns:
        out["成交值燈號"] = ""
    if "成交值排名" in out.columns:
        rk = pd.to_numeric(out["成交值排名"], errors="coerce")
        out["成交值燈號"] = rk.map(lambda x: _emoji_light(_as_float(x), 600, 200, reverse=True))

    # Squeeze pressure unified
    if "嘎空壓力" not in out.columns:
        out["嘎空壓力"] = float("nan")

    smr_col = None
    for c in ["券資比(%)","short_margin_ratio(%)","SMR(%)"]:
        if c in out.columns:
            smr_col = c; break
    if smr_col:
        smr = pd.to_numeric(out[smr_col], errors="coerce")
        sq = pd.to_numeric(out["嘎空壓力"], errors="coerce")
        need = sq.isna() & smr.notna()
        derived = ((smr - 9.0) / (30.0 - 9.0)).clip(lower=0.0, upper=1.0)
        out.loc[need, "嘎空壓力"] = derived.loc[need].round(4)

    # Proxy if still missing
    sq2 = pd.to_numeric(out["嘎空壓力"], errors="coerce")
    need2 = sq2.isna()
    vr = pd.to_numeric(out.get("波動比(20/60)"), errors="coerce")
    qbr = pd.to_numeric(out.get("量比(5/20)"), errors="coerce")
    proxy = ((vr - 1.0).fillna(0.0)*0.5 + (qbr - 1.0).fillna(0.0)*0.5).clip(lower=0.0, upper=1.0)
    out.loc[need2, "嘎空壓力"] = proxy.loc[need2].round(4)

    if "嘎空壓力燈號" not in out.columns:
        out["嘎空壓力燈號"] = ""
    sqv = pd.to_numeric(out["嘎空壓力"], errors="coerce")
    out["嘎空壓力燈號"] = sqv.map(lambda x: _emoji_light(_as_float(x), 0.33, 0.66, reverse=True))
    # NOTE: LOW display is applied only on export copies (to keep numeric for scoring)



    return out




def apply_low_display(df):
    """Return a copy where squeeze pressure shows 'LOW' when SMR < 9% (display-only)."""
    try:
        import pandas as pd
        out = df.copy()
        smr_col2 = None
        for _c in ["券資比(%)", "short_margin_ratio(%)", "SMR(%)"]:
            if _c in out.columns:
                smr_col2 = _c
                break
        if smr_col2 and "嘎空壓力" in out.columns:
            smr2 = pd.to_numeric(out[smr_col2], errors="coerce")
            low = smr2.notna() & (smr2 < 9.0)
            out.loc[low, "嘎空壓力"] = "LOW"
        return out
    except Exception:
        return df

def apply_alignment_and_lights(wb):
    """Set alignment (H=right, V=bottom) and color emoji light columns (燈號)."""
    from openpyxl.styles import Alignment, Font, PatternFill

    color_map = {"🔴": "FFFF0000", "🟡": "FFFFA500", "🟢": "FF00AA00"}
    fill_map  = {"🔴": "FFFFE5E5", "🟡": "FFFFF2CC", "🟢": "FFE2F0D9"}
    na_fill = "FFF2F2F2"

    align = Alignment(horizontal="right", vertical="bottom")
    emoji_align = Alignment(horizontal="center", vertical="bottom")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align

        headers = {}
        for cell in ws[1]:
            if cell.value is None:
                continue
            headers[str(cell.value)] = cell.column

        for h, col in headers.items():
            if "燈號" not in h:
                continue
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                v = c.value
                if v in color_map:
                    c.font = Font(color=color_map[v], bold=True)
                    c.fill = PatternFill("solid", fgColor=fill_map.get(v, na_fill))
                    c.alignment = emoji_align
                elif v in ("N/A", "LOW"):
                    c.font = Font(color="FF666666", bold=True)
                    c.fill = PatternFill("solid", fgColor=na_fill)
                    c.alignment = emoji_align

def postprocess_excel(filepath: str):
    """Post-process an .xlsx file: apply lights, alignment and color."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return
    try:
        wb = load_workbook(filepath)
        apply_alignment_and_lights(wb)
        wb.save(filepath)
    except Exception as e:
        try:
            log(f"[WARN] postprocess_excel failed: {e!r}")
        except Exception:
            pass


def apply_light_formatting(ws, header_row: int = 1):
    """Apply simple font colors for emoji light columns (燈號)."""
    try:
        from openpyxl.styles import Font
    except Exception:
        return
    color_map = {"🔴":"FFFF0000", "🟡":"FFFFA500", "🟢":"FF00AA00"}
    headers = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        headers[str(cell.value)] = cell.column
    for h, col in headers.items():
        if "燈號" not in h:
            continue
        for r in range(header_row+1, ws.max_row+1):
            v = ws.cell(row=r, column=col).value
            if v in color_map:
                ws.cell(row=r, column=col).font = Font(color=color_map[v], bold=True)

def load_json_safe(path: str, default=None):
    try:
        if not os.path.exists(path):
            return default if default is not None else {}
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default if default is not None else {}

def save_json_safe(path: str, obj) -> None:
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False, indent=2)
    except Exception:
        pass



def _parse_int_maybe(x):
    try:
        if x is None:
            return None
        s = str(x).strip().replace(",", "")
        if s in ("", "nan", "None"):
            return None
        return int(float(s))
    except Exception:
        return None


def fetch_shares_outstanding_official_map(timeout=25):
    """Build shares outstanding map from TWSE/TPEx opendata CSV (no Yahoo).

    Sources (daily):
    - Listed:  https://mopsfin.twse.com.tw/opendata/t187ap03_L.csv
    - OTC:     https://mopsfin.twse.com.tw/opendata/t187ap03_O.csv

    Returns:
      dict[str,int] mapping stock code (e.g., '2330') -> issued shares (股數)
    """
    urls = [
        ("TWSE", "https://mopsfin.twse.com.tw/opendata/t187ap03_L.csv"),
        ("TWO",  "https://mopsfin.twse.com.tw/opendata/t187ap03_O.csv"),
    ]
    out = {}
    for tag, url in urls:
        try:
            r = SESSION.get(url, timeout=timeout, headers={"User-Agent": globals().get("UA", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36")})
            r.encoding = "utf-8-sig"
            if r.status_code != 200:
                log(f"official shares map fetch failed ({tag}): http {r.status_code}")
                continue
            # Some hosts return HTML on error; guard by checking header or first chars
            head = (r.text or "")[:200].lower()
            if "<html" in head or "<!doctype" in head:
                log(f"official shares map fetch failed ({tag}): got html")
                continue

            # Read CSV
            from io import StringIO
            df = pd.read_csv(StringIO(r.text))
            # Normalize possible column names
            code_col = None
            shares_col = None

            for c in df.columns:
                if str(c).strip() in ("公司代號", "股票代號", "證券代號", "Security Code", "Company Code"):
                    code_col = c
                if str(c).strip() in ("已發行普通股數或TDR原股發行股數", "已發行普通股數", "已發行股數", "發行股數", "Number of Shares"):
                    shares_col = c
            if code_col is None:
                # try fuzzy
                for c in df.columns:
                    if "代號" in str(c):
                        code_col = c
                        break
            if shares_col is None:
                for c in df.columns:
                    if "已發行" in str(c) and "股" in str(c):
                        shares_col = c
                        break
                if shares_col is None:
                    for c in df.columns:
                        if "發行" in str(c) and "股" in str(c):
                            shares_col = c
                            break

            if code_col is None or shares_col is None:
                log(f"official shares map missing cols ({tag}): cols={list(df.columns)[:15]}")
                continue

            df = df[[code_col, shares_col]].copy()
            df[code_col] = df[code_col].astype(str).str.strip()
            df[shares_col] = df[shares_col].apply(_parse_int_maybe)

            n0 = 0
            for code_, sh in zip(df[code_col].tolist(), df[shares_col].tolist()):
                code_ = str(code_).strip()
                if re.fullmatch(r"\d{4,6}", code_) and sh and sh > 0:
                    out[code_] = int(sh)
                    n0 += 1
            log(f"official shares map loaded ({tag}) n={n0}")
        except Exception as e:
            log(f"official shares map fetch exception ({tag}): {repr(e)}")
    return out


def load_or_build_shares_map(cache_path: str, max_age_hours: int = 72):
    """Cache issued shares map locally; rebuild if missing/old."""
    try:
        import time
        if os.path.exists(cache_path):
            mtime = os.path.getmtime(cache_path)
            age_hours = (time.time() - mtime) / 3600.0
            if age_hours <= max_age_hours:
                data = load_json_safe(cache_path, default={})
                # ensure keys are digits
                data2 = {str(k): int(v) for k, v in data.items() if re.fullmatch(r"\d{4,6}", str(k)) and _parse_int_maybe(v)}
                if data2:
                    return data2
        data = fetch_shares_outstanding_official_map()
        if data:
            save_json_safe(cache_path, data)
        return data
    except Exception:
        return {}

def fetch_shares_outstanding_yahoo_safe(yahoo_symbol: str):
    """Return shares outstanding from Yahoo if possible; otherwise None."""
    try:
        import yfinance as yf
        t = yf.Ticker(yahoo_symbol)
        info = getattr(t, "info", None) or {}
        so = info.get("sharesOutstanding") or info.get("shares_outstanding") or None
        if so is None:
            return None
        so = float(so)
        return so if so > 0 else None
    except Exception:
        return None
import os, re, math, time, json, requests
import os
import pandas as pd
from strategy_score import apply_live_scoring
from lights_unified import apply_lights, apply_display_overrides


def get_latest_volume_from_prices(df_prices: pd.DataFrame):
    """Return latest volume from a single-ticker OHLCV dataframe."""
    try:
        if df_prices is None or len(df_prices) == 0:
            return None
        for col in ["Volume", "volume", "VOL", "成交量"]:
            if col in df_prices.columns:
                v = df_prices[col].iloc[-1]
                try:
                    v = float(v)
                    return v if v >= 0 else None
                except Exception:
                    return None
        return None
    except Exception:
        return None
import math
import json
import csv
import time
import io
# ===== SAFE MODE =====
# --- Column orders (ZH) ---
DECISION_COL_ORDER_ZH = ['進場日期', 'Yahoo代碼', '股票名稱', '策略說明', '進場價', '停損價', '乖離率(%)', '周轉率(%)', '周轉率燈號', '綜合分數', '嘎空壓力', '嘎空壓力燈號', '成交值(元)', '成交值排名', '成交值燈號', '建議部位(元)', '風險提醒']
TOP20_COL_ORDER_ZH = ['進場日期', 'Yahoo代碼', '股票名稱', '策略說明', '進場價', '停損價', '嘎空壓力', '嘎空壓力燈號', '乖離率(%)', '周轉率(%)', '周轉率燈號', '成交值(元)', '成交值排名', '成交值燈號', '建議部位(元)', '風險提醒', '綜合分數']


# --- Fixed FULL column order (ZH) ---
FULL_COL_ORDER_ZH = ['進場日期', 'Yahoo代碼', '股票名稱', '策略代碼', '策略說明', '進場價', '停損價', '收盤價', 'MA20', '乖離率(%)', '周轉率(%)', '周轉率燈號', '一個月支撐', 'ATR20', '波動比(20/60)', '量比(5/20)', '券資比(%)', '嘎空壓力', '權重模式', '權重來源檔', '0', '平滑前權重', '上一期權重', '使用權重', '策略權重(相容)', '平滑係數', '最小樣本數', '樣本筆數', '年化(%)', 'MDD(%)', '建議部位(元)', '風險提醒', '狀態', '出場日期', '出場價', '報酬(%)', '持有天數', '券資比狀態', '券資比判讀', '券資比燈號']


# --- oneclick postprocess shared globals ---
decision_path = None
full_path = None
today = None
market_regime = ""


SAFE_MODE = True  # True=實盤安全; False=嚴格debug

TURNOVER_PRICE_SPLIT = float(os.environ.get("TURNOVER_PRICE_SPLIT", "200"))  # price boundary (NTD)
AMOUNT_PRICE_SPLIT = float(os.environ.get("AMOUNT_PRICE_SPLIT", str(TURNOVER_PRICE_SPLIT)))  # price boundary (NTD)
AMOUNT_BLOCK_NTD_LOW  = float(os.environ.get("AMOUNT_BLOCK_NTD_LOW",  "3000000"))  # <split: min daily traded value (NTD)
AMOUNT_BLOCK_NTD_HIGH = float(os.environ.get("AMOUNT_BLOCK_NTD_HIGH", "6000000"))  # >=split: min daily traded value (NTD)
VOL_TARGET_ANNUAL = float(os.environ.get("VOL_TARGET_ANNUAL", "0.35"))  # target annualized vol for sizing
VOL_LOOKBACK = int(os.environ.get("VOL_LOOKBACK", "20"))
TURNOVER_BLOCK_PCT_LOW  = float(os.environ.get("TURNOVER_BLOCK_PCT_LOW",  "0.20"))  # low price stocks (<split): block if turnover < this (%)
TURNOVER_BLOCK_PCT_HIGH = float(os.environ.get("TURNOVER_BLOCK_PCT_HIGH", "0.10"))  # high price stocks (>=split): block if turnover < this (%)


def turnover_threshold_by_price(price):
    """Return turnover block threshold (%) based on price bucket."""
    try:
        p = float(price)
        if p >= TURNOVER_PRICE_SPLIT:
            return TURNOVER_BLOCK_PCT_HIGH
        return TURNOVER_BLOCK_PCT_LOW
    except Exception:
        # if price missing, use conservative LOW threshold
        return TURNOVER_BLOCK_PCT_LOW


def amount_threshold_by_price(price):
    """Return minimum daily traded value (NTD) based on price bucket."""
    try:
        p = float(price)
        if p >= AMOUNT_PRICE_SPLIT:
            return AMOUNT_BLOCK_NTD_HIGH
        return AMOUNT_BLOCK_NTD_LOW
    except Exception:
        return AMOUNT_BLOCK_NTD_LOW



def ensure_turnover_before_bias(df: pd.DataFrame):
    """Ensure turnover column is present and placed right before bias column when both exist."""
    try:
        tcol = None
        for c in ["turnover_rate(%)", "周轉率(%)"]:
            if c in df.columns:
                tcol = c
                break
        bcol = None
        for c in ["bias(%)", "乖離率(%)"]:
            if c in df.columns:
                bcol = c
                break
        if tcol and bcol:
            cols = [c for c in df.columns if c != tcol]
            bidx = cols.index(bcol)
            cols.insert(bidx, tcol)
            return df.reindex(columns=cols)
        return df
    except Exception:
        return df
from utils_safe import safe_get, safe_float, safe_round
# ======================
def normalize_market_code(market: str) -> str:
    """
    將市場欄位統一為內部代碼：TWSE / TWO
    允許輸入：TWSE/TW/上市、TWO/OTC/上櫃 等。
    """
    s = str(market).strip()
    su = s.upper()
    if su in ("TW", "TWSE", "LISTED") or s == "上市":
        return "TWSE"
    if su in ("TWO", "OTC", "TPEX") or s == "上櫃":
        return "TWO"
    return su


def vol_scale_factor(vol_annual):
    """Scale factor for position sizing based on annualized volatility."""
    try:
        v = float(vol_annual)
        if v <= 0 or pd.isna(v):
            return 1.0
        return float(min(1.0, max(0.0, VOL_TARGET_ANNUAL / v)))
    except Exception:
        return 1.0



def format_smr_display(val, meta: str = ""):
    """券資比顯示：None/NaN/0→N/A；DIV0→∞；其餘→小數點2位"""
    m = str(meta).upper().strip()
    if m in ("DIV0", "INF"):
        return "∞"
    if val is None:
        return "N/A"
    try:
        f = float(val)
        if math.isnan(f) or f == 0.0:
            return "N/A"
        return f"{f:.2f}"
    except Exception:
        return "N/A"


def smr_text_label(val, meta: str = "", high30: float = 30.0, block50: float = 50.0):
    m = str(meta).upper().strip()
    if m in ("DIV0", "INF"):
        return "禁止(除0)"
    if m in ("NO_DATA", "NA"):
        return "N/A"
    try:
        if val is None:
            return "N/A"
        f = float(val)
        if math.isnan(f) or f == 0.0:
            return "N/A"
    except Exception:
        return "N/A"
    if f >= block50:
        return "禁止(>=50)"
    if f >= high30:
        return "偏高(>=30)"
    if f < 10:
        return "低(<10)"
    return "正常"




def _shares_cache_path():
    return os.path.join("daily_excel_records", "shares_cache.json")

def _load_shares_cache():
    path = _shares_cache_path()
    try:
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_shares_cache(cache: dict):
    path = _shares_cache_path()
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def fetch_shares_outstanding(tickers: list[str], safe_mode: bool = True) -> dict:
    """只針對入選股票抓 sharesOutstanding（慢），並寫入本地 cache。"""
    cache = _load_shares_cache()
    out = {}
    need = []
    for t in tickers:
        if t in cache:
            try:
                out[t] = int(cache[t])
                continue
            except Exception:
                pass
        need.append(t)

    if not need:
        return out

    try:
        import yfinance as yf
    except Exception:
        return out

    for t in need:
        try:
            info = yf.Ticker(t).info
            so = info.get("sharesOutstanding") or info.get("shares_outstanding")
            if so:
                so_int = int(so)
                out[t] = so_int
                cache[t] = so_int
        except Exception:
            if not safe_mode:
                raise
        time.sleep(0.2)

    _save_shares_cache(cache)
    return out

def smr_traffic_light(label: str) -> str:
    s = str(label)
    if s.startswith("禁止"):
        return "紅"
    if s.startswith("偏高"):
        return "黃"
    if s in ("低(<10)", "正常"):
        return "綠"
    return "灰"

import yfinance as yf
from io import StringIO
from datetime import datetime, timedelta
import smtplib
from email.message import EmailMessage

import logging
logging.getLogger("yfinance").setLevel(logging.CRITICAL)  # v6.3.1


# -------- runtime logging (v6.2.2) --------
RUN_TS = datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE = f"run_{RUN_TS}.log"
INVALID_TICKERS_FILE = f"invalid_tickers_{RUN_TS}.csv"
INVALID_TICKERS = set()
NAME_MAP = {}
STRAT_STATS = {}  # strategy -> dict(annualized_pct, mdd_pct, trades_used)

def log(msg: str) -> None:
    s = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(s, flush=True)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(s + "\n")
    except Exception:
        pass
# ------------------------------------------


def set_name_map(isin_df: pd.DataFrame) -> None:
    """
    v6.3.12: 建立 (symbol, market) -> name_zh 的查表，避免 merge 或 meta 丟失造成名稱空白。
    """
    global NAME_MAP
    try:
        if isin_df is None or isin_df.empty:
            return
        df = isin_df.copy()
        if "name_zh" not in df.columns:
            df["name_zh"] = ""
        d = {}
        for _, r in df[["symbol","market","name_zh"]].dropna(subset=["symbol","market"]).iterrows():
            d[(str(r["symbol"]).strip(), str(r["market"]).strip())] = str(r["name_zh"]).strip()
        NAME_MAP = d
        log(f"NAME_MAP built: {len(NAME_MAP)}")
    except Exception as e:
        log("NAME_MAP build failed: " + repr(e))

def flush_invalid_tickers() -> None:
    try:
        if not INVALID_TICKERS:
            return
        import pandas as _pd
        _pd.DataFrame({"ticker": sorted(INVALID_TICKERS)}).to_csv(INVALID_TICKERS_FILE, index=False, encoding="utf-8-sig")
        log(f"Wrote invalid tickers: {INVALID_TICKERS_FILE} (n={len(INVALID_TICKERS)})")
    except Exception:
        pass




TOTAL_CAPITAL = 300_000
MAX_ACCOUNT_RISK = 0.10
PER_TRADE_RISK_RATIO = 0.20
MIN_AVG_VOLUME = 100_000

# ===== 兩階段快篩加速（v6.3）=====
ENABLE_TWO_STAGE_SCREEN = True
STAGE1_PERIOD = "5d"          # 全市場只抓近 5 天（很快）
STAGE2_PERIOD = "6mo"         # 只對候選抓完整歷史
TOPN_LIQUID = 1200             # 取成交量前 N 名進入 Stage2
# =================================

LOCAL_EXCEL_FOLDER = "daily_excel_records"
SENDER_EMAIL = "fish262829@gmail.com"
APP_PASSWORD_ENV = "GMAIL_APP_PASSWORD"
RECEIVER_EMAIL = "fish262829@gmail.com"

# 是否寄送 Email（v6.2.4）
ENABLE_EMAIL = False  # True=寄信；False=只產出Excel

INDEX_TICKER = "0050.TW"
PERF_SUMMARY_FILE = "performance_summary.xlsx"

BATCH_SIZE = 60  # v6.3.3 reduce to avoid Yahoo rate limit
CACHE_DIR = "cache_yf"
CACHE_TTL_DAYS = 2

# v6.3.3: yfinance 節流參數（避免 RateLimit）
YF_THREADS = False
SLEEP_BETWEEN_YF_BATCH = 0.8
YF_MAX_RETRY = 4
YF_BACKOFF_BASE = 3.0

TH_BIAS_MEAN_REVERT = -6.0
TH_BIAS_DEEP = -8.0
TH_SQUEEZE_TW = 30.0
TH_SQUEEZE_OTC_PROXY = 0.90
TH_SUPPORT_TOL = 0.00
TH_SQUEEZE_VOL_RATIO = 0.70
TH_SQUEEZE_VOLRATIO_VOL = 1.20
TH_SQUEEZE_MIN_BIAS = -3.0
TH_HIGH_RISK_POS = 0.20

REGIME_RISK_MULTIPLIER = {"RANGE": 1.0, "TREND": 1.2, "STRESS": 0.5}

ENABLE_WEIGHT_SMOOTHING = True
WEIGHT_SMOOTH_ALPHA = 0.30
MIN_TRADES_FOR_WEIGHT = 20
WEIGHT_STATE_FILE = "strategy_weights_state.json"
MAX_STRATEGY_WEIGHT = 0.50

# ===== Excel 中文欄位顯示（v6.2.4）=====
# ===== Excel 中文欄位顯示（v6.3.5 最終實盤版）=====
# 重要：dict 會決定輸出欄位順序
# ===== Excel 中文欄位顯示（v6.3.6 最終實盤版）=====
# 重要：dict 會決定輸出欄位順序
# ===== Excel 中文欄位顯示（v6.3.8 最終實盤版）=====
# 重要：dict 會決定輸出欄位順序
COLUMN_MAP_ZH = {
    "turnover_light": "周轉率燈號",

    "entry_date": "進場日期",
    "ticker": "Yahoo代碼",
    "name_zh": "股票名稱",
    "symbol": "股票代號",
    "market": "市場",
    "strategy": "策略代碼",
    "strategy_desc": "策略說明",
    "entry_price": "進場價",
    "stop_loss_price": "停損價",
    "close": "收盤價",
    "ma20": "MA20",
    "turnover_rate(%)": "周轉率(%)",
    "bias20": "乖離率(%)",
    "support_1m": "一個月支撐",
    "atr20": "ATR20",
    "volatility_ratio": "波動比(20/60)",
    "volume_ratio": "量比(5/20)",
    "short_margin_ratio(%)": "券資比(%)",

    "otc_short_pressure": "嘎空壓力",
    "weight_mode": "權重模式",
    "weight_source_file": "權重來源檔",
    "weight_score_raw": (safe_float(safe_get(locals(), "weight_score_raw", 0.0, SAFE_MODE), 0.0, SAFE_MODE) + safe_float(safe_get(locals(), "score_penalty", 0.0, SAFE_MODE), 0.0, SAFE_MODE)),
    "weight_raw": "平滑前權重",
    "weight_prev": "上一期權重",
    "weight_used": "使用權重",
    "strategy_weight": "策略權重(相容)",
    "weight_alpha": "平滑係數",
    "min_trades": "最小樣本數",
    "trades_used": "樣本筆數",
    "annualized_pct": "年化(%)",
    "mdd_pct": "MDD(%)",
    "position_size": "建議部位(元)",
    "Risk Alert": "風險提醒",
    "status": "狀態",
    "exit_date": "出場日期",
    "exit_price": "出場價",
    "pnl_pct": "報酬(%)",
    "hold_days": "持有天數",
    "smr_status": "券資比狀態",
    "SMR Label": "券資比判讀",
    "SMR Light": "券資比燈號",
}
COLUMN_MAP_EN = {v: k for k, v in COLUMN_MAP_ZH.items()}
# =======================================
# =======================================
# =======================================

# ===== 策略中文說明（顯示用）=====
STRATEGY_DESC_ZH = {
    "SQUEEZE_TW": "上市嘎空：券資比高 + 波動/量能放大 + 價格站上MA20",
    "SQUEEZE_OTC": "上櫃嘎空代理：波動/量能轉強（無券資比時以壓力指標替代）",
    "MEAN_REVERT": "乖離回歸：BIAS(20) <= 門檻，靠近支撐",
    "HIGH_MARGIN_MEAN_REVERT": "高融資回歸：更深度乖離（風險較高）",
}
# =================================
# =======================================


SESSION = requests.Session()

# v6.3.21.0: meta for margin/short ratio availability
MARGIN_RATIO_META = {}  # (symbol, market)->status

SESSION.headers.update({"User-Agent": "Mozilla/5.0"})


def _fetch_isin_universe(str_mode: int) -> pd.DataFrame:
    """
    ISIN 名單（HTML）取得股票代號 + 中文名稱
    - strMode=2: 上市
    - strMode=4: 上櫃

    內部欄位：symbol / market / name_zh
    v6.3.10: 強制回傳 name_zh 欄；並記錄非空筆數。
    """
    url = "https://isin.twse.com.tw/isin/C_public.jsp"
    r = SESSION.get(url, params={"strMode": str_mode}, timeout=25)
    r.encoding = "big5"

    html = r.text or ""
    start = html.find("<table")
    end = html.rfind("</table>")
    html_snip = html[start:end + len("</table>")] if (start != -1 and end != -1 and end > start) else html

    market = "TW" if str_mode == 2 else "TWO"

    def _final(df_sym_name: pd.DataFrame) -> pd.DataFrame:
        if "name_zh" not in df_sym_name.columns:
            df_sym_name["name_zh"] = ""
        df_sym_name["symbol"] = df_sym_name["symbol"].astype(str).str.strip()
        df_sym_name = df_sym_name[df_sym_name["symbol"].str.fullmatch(r"\d{4,6}")].copy()
        df_sym_name["market"] = market
        df_sym_name["name_zh"] = df_sym_name["name_zh"].astype(str).str.strip()
        df_sym_name.loc[df_sym_name["name_zh"].isin(["nan", "None"]), "name_zh"] = ""
        return df_sym_name[["symbol", "market", "name_zh"]].reset_index(drop=True)

    try:
        from io import StringIO
        tables = pd.read_html(StringIO(html_snip))
        if tables:
            df0 = tables[0].copy()
            first_col = df0.columns[0]
            series = df0[first_col].astype(str).str.strip()
            parts = series.str.split(r"[\s\u3000]+", n=1, expand=True)
            sym = parts[0]
            name = parts[1] if parts.shape[1] > 1 else ""
            out = _final(pd.DataFrame({"symbol": sym, "name_zh": name}).dropna(subset=["symbol"]))
            non_empty = int((out["name_zh"].astype(str).str.len() > 0).sum()) if "name_zh" in out.columns else 0
            if not out.empty:
                log(f"ISIN success: strMode={str_mode}, n={len(out)}, name_nonempty={non_empty}")
                return out
    except Exception as e:
        log(f"ISIN read_html failed: strMode={str_mode}, err={repr(e)}")

    codes = sorted(set(re.findall(r"(?<!\d)(\d{4,6})(?!\d)", html_snip)))
    out = _final(pd.DataFrame({"symbol": codes, "name_zh": [""] * len(codes)}))
    log(f"ISIN regex fallback used: strMode={str_mode}, n={len(out)}")
    return out

def _last_business_day(dt: datetime) -> datetime:
    d = dt
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d

def _yyyymmdd(dt: datetime) -> str:
    return dt.strftime("%Y%m%d")

def _today_str() -> str:
    return _last_business_day(_dt.datetime.now()).strftime("%Y-%m-%d")


def fetch_listed_stocks() -> pd.DataFrame:
    """
    Primary: TWSE STOCK_DAY_ALL (CSV)
    Fallback: TWSE ISIN list (strMode=2) when CSV fails / format changes.
    """
    url = "https://www.twse.com.tw/exchangeReport/STOCK_DAY_ALL"
    try:
        r = SESSION.get(url, params={"response": "csv"}, timeout=20)
        r.encoding = "utf-8"
        df = pd.read_csv(StringIO(r.text), header=1)
        df = df[["證券代號"]].dropna()
        df["market"] = "TW"
        df = df.rename(columns={"證券代號": "symbol"})
        df["symbol"] = df["symbol"].astype(str).str.strip()
        df = df[df["symbol"].str.fullmatch(r"\d+")]
        if not df.empty:
            return df.reset_index(drop=True)
    except Exception:
        pass
    return _fetch_isin_universe(str_mode=2)

def fetch_otc_stocks() -> pd.DataFrame:
    """
    Primary: TPEx OpenAPI (JSON)
    Fallback: TWSE ISIN list (strMode=4) when OpenAPI returns HTML / empty / blocked.
    """
    url = "https://www.tpex.org.tw/openapi/v1/mopsfin_t187ap03_L"
    try:
        r = SESSION.get(url, timeout=20, headers={"Accept": "application/json"})
        ctype = (r.headers.get("Content-Type", "") or "").lower()
        if r.status_code == 200 and "application/json" in ctype:
            j = r.json()
            df = pd.DataFrame(j)
            df = df[["SecuritiesCompanyCode"]].dropna()
            df["market"] = "TWO"
            df = df.rename(columns={"SecuritiesCompanyCode": "symbol"})
            df["symbol"] = df["symbol"].astype(str).str.strip()
            df = df[df["symbol"].str.fullmatch(r"\d+")]
            if not df.empty:
                return df.reset_index(drop=True)
    except Exception:
        pass
    return _fetch_isin_universe(str_mode=4)

def _parse_twse_csv_loose(text: str) -> pd.DataFrame:
    try:
        return pd.read_csv(StringIO(text), header=1)
    except Exception:
        lines = []
        for ln in text.splitlines():
            s = ln.strip()
            if re.match(r"^\d{4,6},", s):
                lines.append(s)
        if not lines:
            return pd.DataFrame()
        return pd.read_csv(StringIO("\n".join(lines)), header=None)

def fetch_twse_short_margin_ratio(latest_dt: datetime, max_lookback_days: int = 10) -> pd.DataFrame:
    url = "https://www.twse.com.tw/exchangeReport/MI_MARGN"
    for i in range(max_lookback_days):
        dt = _last_business_day(latest_dt - timedelta(days=i))
        date_str = _yyyymmdd(dt)
        try:
            r = SESSION.get(url, params={"response": "csv", "date": date_str, "selectType": "ALL"}, timeout=25)
            r.encoding = "utf-8"
            df0 = _parse_twse_csv_loose(r.text)
            if df0.empty:
                continue
            col_map = {}
            for c in df0.columns:
                cs = str(c)
                if ("股票代號" in cs) or ("證券代號" in cs):
                    col_map[c] = "symbol"
                if ("融資" in cs) and ("餘額" in cs):
                    col_map[c] = "margin"
                if ("融券" in cs) and ("餘額" in cs):
                    col_map[c] = "short"
            df = df0.rename(columns=col_map)
            if not {"symbol", "margin", "short"}.issubset(set(df.columns)):
                continue
            df = df[["symbol", "margin", "short"]].copy()
            df["symbol"] = df["symbol"].astype(str).str.strip()
            df = df[df["symbol"].str.fullmatch(r"\d+")]
            for c in ["margin", "short"]:
                df[c] = pd.to_numeric(df[c].astype(str).str.replace(",", ""), errors="coerce")
            df = df.dropna(subset=["margin", "short"])
            df = df[df["margin"] > 0]
            df["short_margin_ratio"] = (df["short"] / df["margin"]) * 100.0
            out = df[["symbol", "short_margin_ratio"]].copy()
            out = out.replace([math.inf, -math.inf], pd.NA).dropna(subset=["short_margin_ratio"])
            if not out.empty:
                print(f"TWSE MI_MARGN date used: {date_str} (T+1 data)")
                return out.reset_index(drop=True)
        except Exception:
            time.sleep(0.2)
            continue
    return pd.DataFrame(columns=["symbol", "short_margin_ratio"])


def _cache_path_for_ticker(ticker: str) -> str:
    safe = ticker.replace("^", "IDX_")
    return os.path.join(CACHE_DIR, f"{safe}.pkl")

def _cache_is_fresh(path: str) -> bool:
    try:
        mtime = datetime.fromtimestamp(os.path.getmtime(path))
        return (datetime.now() - mtime).days <= CACHE_TTL_DAYS
    except Exception:
        return False

def load_cached_history(ticker: str):
    path = _cache_path_for_ticker(ticker)
    if os.path.exists(path) and _cache_is_fresh(path):
        try:
            return pd.read_pickle(path)
        except Exception:
            return None
    return None

def save_cached_history(ticker: str, df: pd.DataFrame) -> None:
    os.makedirs(CACHE_DIR, exist_ok=True)
    ratio_map = {}
    try:
        df.to_pickle(_cache_path_for_ticker(ticker))
    except Exception:
        pass


def yf_download_with_retry(tickers: list[str], period: str) -> pd.DataFrame:
    """
    v6.3.3: yfinance 下載加入 RateLimit 重試 + backoff
    """
    last_err = None
    for attempt in range(1, YF_MAX_RETRY + 1):
        try:
            return yf.download(
                " ".join(tickers),
                period=period,
                group_by="ticker",
                threads=YF_THREADS,
                auto_adjust=False,
                progress=False,
            )
        except Exception as e:
            msg = str(e)
            last_err = e
            if ("Too Many Requests" in msg) or ("Rate limited" in msg) or ("YFRateLimitError" in msg):
                sleep_s = YF_BACKOFF_BASE * attempt
                log(f"yfinance rate limited. retry {attempt}/{YF_MAX_RETRY} sleep {sleep_s:.1f}s")
                time.sleep(sleep_s)
                continue
            raise
    raise RuntimeError(f"yfinance download failed after retries: {repr(last_err)}")

def batched(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i+n]



# ===== v6.3.17: 融資融券 / 券資比 (T+1) =====
def _fmt_date_yyyymmdd(d: _dt.date) -> str:
    return d.strftime("%Y%m%d")

def _fmt_date_roc_slash(d: _dt.date) -> str:
    roc_year = d.year - 1911
    return f"{roc_year}/{d.month:02d}/{d.day:02d}"

def _try_fetch_twse_margin(date_yyyymmdd: str) -> pd.DataFrame:
    """
    TWSE MI_MARGN (ALL) -> DataFrame with columns: symbol, margin_balance, short_balance, market
    """
    url = "https://www.twse.com.tw/exchangeReport/MI_MARGN"
    r = SESSION.get(url, params={"response": "json", "date": date_yyyymmdd, "selectType": "ALL"}, timeout=25)
    try:
        j = r.json()
    except Exception:
        return pd.DataFrame()
    data = j.get("data") or []
    fields = j.get("fields") or []
    if not data or not fields:
        return pd.DataFrame()
    df = pd.DataFrame(data, columns=fields)

    col_code = None
    for c in df.columns:
        s = str(c)
        if "證券代號" in s or "股票代號" in s:
            col_code = c
            break
    if col_code is None:
        col_code = df.columns[0]

    def _find_col(kw: str):
        for c in df.columns:
            if kw in str(c):
                return c
        return None

    c_margin = _find_col("融資餘額")
    c_short = _find_col("融券餘額")
    if c_margin is None or c_short is None:
        return pd.DataFrame()

    out = pd.DataFrame({
        "symbol": df[col_code].astype(str).str.strip(),
        "margin_balance": pd.to_numeric(df[c_margin].astype(str).str.replace(",", ""), errors="coerce"),
        "short_balance": pd.to_numeric(df[c_short].astype(str).str.replace(",", ""), errors="coerce"),
        "market": "TWSE",
    })
    out = out[out["symbol"].str.fullmatch(r"\d{4,6}")].copy()
    return out

def _try_fetch_tpex_margin(date_roc_slash: str) -> pd.DataFrame:
    """
    TPEx margin balance endpoint (JSON) -> DataFrame with columns: symbol, margin_balance, short_balance, market
    """
    url = "https://www.tpex.org.tw/web/stock/margin_trading/margin_balance/margin_bal_result.php"
    r = SESSION.get(url, params={"l": "zh-tw", "o": "json", "d": date_roc_slash, "s": "0,asc"}, timeout=25)
    try:
        j = r.json()
    except Exception:
        return pd.DataFrame()
    data = j.get("aaData") or j.get("data") or []
    if not data:
        return pd.DataFrame()
    df = pd.DataFrame(data)
    if df.shape[1] < 5:
        return pd.DataFrame()
    sym = df.iloc[:, 0].astype(str).str.strip()
    margin = pd.to_numeric(df.iloc[:, 2].astype(str).str.replace(",", ""), errors="coerce")
    short = pd.to_numeric(df.iloc[:, 4].astype(str).str.replace(",", ""), errors="coerce")
    out = pd.DataFrame({"symbol": sym, "margin_balance": margin, "short_balance": short, "market": "TWO"})
    out = out[out["symbol"].str.fullmatch(r"\d{4,6}")].copy()
    return out

def _req_json(url: str, params: dict, timeout: int = 25, tries: int = 3) -> dict:
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json,text/plain,*/*",
        "Referer": "https://www.twse.com.tw/",
        "Connection": "keep-alive",
    }
    last = {}
    for _ in range(tries):
        try:
            r = SESSION.get(url, params=params, headers=headers, timeout=timeout)
            try:
                return r.json()
            except Exception:
                last = {"_raw": r.text[:200]}
        except Exception as e:
            last = {"_err": repr(e)}
    return last

def _twse_margn_df(date_yyyymmdd: str) -> pd.DataFrame:
    url = "https://www.twse.com.tw/exchangeReport/MI_MARGN"
    j = _req_json(url, {"response":"json","date":date_yyyymmdd,"selectType":"ALL"})
    data = j.get("data") if isinstance(j, dict) else None
    fields = j.get("fields") if isinstance(j, dict) else None
    stat = (j.get("stat") if isinstance(j, dict) else "") or ""
    if data and fields and ("OK" in stat or stat == "" or "success" in str(stat).lower()):
        df = pd.DataFrame(data, columns=fields)
    else:
        # CSV fallback
        headers = {"User-Agent":"Mozilla/5.0","Referer":"https://www.twse.com.tw/"}
        try:
            r = SESSION.get(url, params={"response":"csv","date":date_yyyymmdd,"selectType":"ALL"}, headers=headers, timeout=25)
            txt = r.text
            lines = [ln for ln in txt.splitlines() if ln.count(",") >= 5 and not ln.startswith("=")]
            if not lines:
                return pd.DataFrame()
            header_idx = None
            for i,ln in enumerate(lines[:40]):
                if "證券代號" in ln and "融資餘額" in ln and "融券餘額" in ln:
                    header_idx = i
                    break
            if header_idx is None:
                return pd.DataFrame()
            import io
            df = pd.read_csv(io.StringIO("\n".join(lines[header_idx:])))
        except Exception:
            return pd.DataFrame()

    col_code = None
    for c in df.columns:
        s=str(c)
        if "證券代號" in s or "股票代號" in s:
            col_code=c; break
    if col_code is None:
        col_code=df.columns[0]

    def _find(kw):
        for c in df.columns:
            if kw in str(c):
                return c
        return None
    c_margin=_find("融資餘額")
    c_short=_find("融券餘額")
    if c_margin is None or c_short is None:
        return pd.DataFrame()

    out = pd.DataFrame({
        "symbol": df[col_code].astype(str).str.strip(),
        "margin_balance": pd.to_numeric(df[c_margin].astype(str).str.replace(",",""), errors="coerce"),
        "short_balance": pd.to_numeric(df[c_short].astype(str).str.replace(",",""), errors="coerce"),
        "market": "TWSE",
    })
    out = out[out["symbol"].str.fullmatch(r"\d{4}")].copy()
    return out

def _tpex_margin_df(date_roc_slash: str) -> pd.DataFrame:
    url = "https://www.tpex.org.tw/web/stock/margin_trading/margin_balance/margin_bal_result.php"
    headers = {"User-Agent":"Mozilla/5.0","Accept":"application/json,text/plain,*/*","Referer":"https://www.tpex.org.tw/"}
    for _ in range(3):
        try:
            r = SESSION.get(url, params={"l":"zh-tw","o":"json","d":date_roc_slash,"s":"0,asc"}, headers=headers, timeout=25)
            j = r.json()
            data = j.get("aaData") or j.get("data") or []
            if not data:
                continue
            df = pd.DataFrame(data)
            if df.shape[1] < 5:
                continue
            sym = df.iloc[:,0].astype(str).str.strip()
            margin = pd.to_numeric(df.iloc[:,2].astype(str).str.replace(",",""), errors="coerce")
            short = pd.to_numeric(df.iloc[:,4].astype(str).str.replace(",",""), errors="coerce")
            out = pd.DataFrame({"symbol": sym, "margin_balance": margin, "short_balance": short, "market":"TWO"})
            out = out[out["symbol"].str.fullmatch(r"\d{4}")].copy()
            return out
        except Exception:
            continue
    return pd.DataFrame()


def _tpex_margin_latest_html_df() -> pd.DataFrame:
    """
    上櫃(TWO) 融資融券餘額（最新一筆）

    優先來源：TPEx OpenAPI v1
      - https://www.tpex.org.tw/openapi/v1/tpex_mainboard_margin_balance
      (此端點通常不需日期，直接回傳最新資料)

    備援來源：TPEx 網頁表格（若 OpenAPI 失效）
      - https://www.tpex.org.tw/zh-tw/mainboard/trading/margin-trading/transactions.html

    產出欄位：symbol, margin_balance, short_balance, market="TWO"
    """
    # ---- 1) OpenAPI v1 (preferred) ----
    try:
        url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_margin_balance"
        headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json,*/*", "Connection": "keep-alive"}
        r = SESSION.get(url, headers=headers, timeout=25)
        r.raise_for_status()
        js = r.json()
        if isinstance(js, dict) and "data" in js:
            js = js["data"]
        if isinstance(js, list) and len(js) > 0 and isinstance(js[0], dict):
            df = pd.DataFrame(js)
            # flatten columns
            df.columns = [str(c).strip() for c in df.columns]

            def pick_col_contains(must_subs):
                for c in df.columns:
                    s = str(c)
                    ok = True
                    for sub in must_subs:
                        if sub not in s:
                            ok = False
                            break
                    if ok:
                        return c
                return None

            # Try common chinese names first
            c_code = pick_col_contains(["代號"]) or pick_col_contains(["股票代號"]) or pick_col_contains(["證券代號"]) or pick_col_contains(["Code"]) or pick_col_contains(["symbol"])
            # "資餘額" or "融資今日餘額"/"融資餘額"
            c_mbal = pick_col_contains(["資", "餘額"]) or pick_col_contains(["融資", "餘額"]) or pick_col_contains(["Margin", "Balance"]) or pick_col_contains(["margin"])
            # "券餘額" or "融券今日餘額"/"融券餘額"
            c_sbal = pick_col_contains(["券", "餘額"]) or pick_col_contains(["融券", "餘額"]) or pick_col_contains(["Short", "Balance"]) or pick_col_contains(["short"])

            if c_code and c_mbal and c_sbal:
                def to_num(x):
                    ss = str(x).strip().replace(",", "")
                    if ss in ("", "-", "—", "–", "nan", "None"):
                        return float("nan")
                    ss = re.sub(r"[^0-9\.\-]", "", ss)
                    try:
                        return float(ss) if ss != "" else float("nan")
                    except Exception:
                        return float("nan")

                out = pd.DataFrame({
                    "symbol": df[c_code].astype(str).str.strip(),
                    "margin_balance": df[c_mbal].apply(to_num),
                    "short_balance": df[c_sbal].apply(to_num),
                    "market": "TWO",
                })
                out = out[out["symbol"].str.fullmatch(r"\d{4}")].copy()
                log(f"TPEx OpenAPI margin_balance parsed rows={len(out)}")
                return out
            else:
                log(f"TPEx OpenAPI missing cols code={c_code} mbal={c_mbal} sbal={c_sbal} cols={df.columns.tolist()[:20]}")
        else:
            log("TPEx OpenAPI returned empty/non-list payload")
    except Exception as e:
        log(f"TPEx OpenAPI parse failed: {repr(e)}")

    # ---- 2) HTML fallback ----
    url = "https://www.tpex.org.tw/zh-tw/mainboard/trading/margin-trading/transactions.html"
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "text/html; charset=utf-8,*/*", "Connection": "keep-alive"}
    try:
        r = SESSION.get(url, headers=headers, timeout=25)
        r.raise_for_status()
        # Some pandas versions treat raw string as path; use io.StringIO
        try:
            r.encoding = r.apparent_encoding or "utf-8"
        except Exception:
            pass
        tables = pd.read_html(io.StringIO(r.text), flavor="lxml")
        if not tables:
            log("TPEx latest HTML: no tables")
            return pd.DataFrame()

        df = tables[0].copy()
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = ["".join([str(x).strip() for x in col if str(x).strip() not in ("nan", "None")]).strip() for col in df.columns.values]
        else:
            df.columns = [str(c).strip() for c in df.columns]

        def find_col(need_substr):
            for c in df.columns:
                s = str(c).strip()
                ok = True
                for sub in need_substr:
                    if sub not in s:
                        ok = False
                        break
                if ok:
                    return c
            return None

        c_code = find_col(["代號"]) or find_col(["證券代號"])
        c_mbal = find_col(["資", "餘額"]) or find_col(["融資", "餘額"])
        c_sbal = find_col(["券", "餘額"]) or find_col(["融券", "餘額"])
        if not (c_code and c_mbal and c_sbal):
            log(f"TPEx latest HTML: missing cols code={c_code} mbal={c_mbal} sbal={c_sbal} cols={df.columns.tolist()[:12]}")
            return pd.DataFrame()

        def to_num(x):
            ss = str(x).strip().replace(",", "")
            if ss in ("", "-", "—", "–", "nan", "None"):
                return float("nan")
            ss = re.sub(r"[^0-9\.\-]", "", ss)
            try:
                return float(ss) if ss != "" else float("nan")
            except Exception:
                return float("nan")

        out = pd.DataFrame({
            "symbol": df[c_code].astype(str).str.strip(),
            "margin_balance": df[c_mbal].apply(to_num),
            "short_balance": df[c_sbal].apply(to_num),
            "market": "TWO",
        })
        out = out[out["symbol"].str.fullmatch(r"\d{4}")].copy()
        log(f"TPEx latest HTML parsed rows={len(out)}")
        return out
    except Exception as e:
        log(f"TPEx latest HTML parse failed: {repr(e)}")
        return pd.DataFrame()

        df = tables[0].copy()

        # flatten multiindex columns
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = ["".join([str(x).strip() for x in col if str(x).strip() not in ("nan","None")]).strip() for col in df.columns.values]
        else:
            df.columns = [str(c).strip() for c in df.columns]

        def find_col(need_substr: list[str]):
            for c in df.columns:
                s = str(c).strip()
                ok = True
                for sub in need_substr:
                    if sub not in s:
                        ok = False
                        break
                if ok:
                    return c
            return None

        # Code column: contains 代號
        c_code = find_col(["代號"]) or find_col(["股票代號"]) or find_col(["證券代號"])
        # Margin balance column: contains 資 and 餘額
        c_mbal = find_col(["資", "餘額"]) or find_col(["融資", "餘額"])
        # Short balance column: contains 券 and 餘額
        c_sbal = find_col(["券", "餘額"]) or find_col(["融券", "餘額"])

        if not (c_code and c_mbal and c_sbal):
            log(f"TPEx latest HTML: missing cols code={c_code} mbal={c_mbal} sbal={c_sbal} cols={df.columns.tolist()[:12]}")
            return pd.DataFrame()

        def to_num(s):
            # remove commas and non-numeric (keep minus and dot)
            ss = str(s).strip().replace(",", "")
            if ss in ("", "-", "—", "–", "nan", "None"):
                return float("nan")
            ss = re.sub(r"[^0-9\.\-]", "", ss)
            try:
                return float(ss) if ss != "" else float("nan")
            except Exception:
                return float("nan")

        out = pd.DataFrame({
            "symbol": df[c_code].astype(str).str.strip(),
            "margin_balance": df[c_mbal].apply(to_num),
            "short_balance": df[c_sbal].apply(to_num),
            "market": "TWO",
        })
        out = out[out["symbol"].str.fullmatch(r"\d{4}")].copy()
        log(f"TPEx latest HTML parsed rows={len(out)} numeric_rows={int(out['margin_balance'].notna().sum())}/{int(out['short_balance'].notna().sum())}")
        return out

    except Exception as e:
        log(f"TPEx latest HTML parse failed: {repr(e)}")
        return pd.DataFrame()
        df = tables[0].copy()
        df.columns = [str(c).strip() for c in df.columns]
        def find_col(cands):
            for c in df.columns:
                s = str(c).strip()
                if s in cands:
                    return c
            for c in df.columns:
                s = str(c).strip()
                for cand in cands:
                    if cand in s:
                        return c
            return None
        c_code = find_col(["代號", "股票代號", "證券代號"])
        c_mbal = find_col(["資餘額", "融資餘額", "資餘額(張)"])
        c_sbal = find_col(["券餘額", "融券餘額", "券餘額(張)"])
        if not (c_code and c_mbal and c_sbal):
            return pd.DataFrame()
        out = pd.DataFrame({
            "symbol": df[c_code].astype(str).str.strip(),
            "margin_balance": pd.to_numeric(df[c_mbal].astype(str).str.replace(",", ""), errors="coerce"),
            "short_balance": pd.to_numeric(df[c_sbal].astype(str).str.replace(",", ""), errors="coerce"),
            "market": "TWO",
        })
        out = out[out["symbol"].str.fullmatch(r"\d{4}")].copy()
        return out
    except Exception as e:
        log(f"TPEx latest HTML parse failed: {repr(e)}")
        return pd.DataFrame()

def _twse_openapi_mi_margn_df() -> pd.DataFrame:
    """
    TWSE OpenAPI v1（通常為最新一筆，不帶日期參數）
    GET https://openapi.twse.com.tw/v1/exchangeReport/MI_MARGN

    注意：此端點有時不直接提供「融券今日餘額」欄，只有「融券前日餘額」與買進/賣出/償還。
    本函式會在缺少「融券今日餘額」時，用以下方式推導：
        融券今日餘額 = 融券前日餘額 + 融券賣出 - 融券買進 - 融券現券償還
    """
    url = "https://openapi.twse.com.tw/v1/exchangeReport/MI_MARGN"
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json,text/plain,*/*",
        "Connection": "keep-alive",
    }
    verify_ssl = (os.getenv("TWSE_SSL_VERIFY", "1").strip() != "0")

    def _find_col(df: pd.DataFrame, cands: list[str]):
        for c in df.columns:
            s = str(c).strip()
            if s in cands:
                return c
        for c in df.columns:
            s = str(c).strip()
            for cand in cands:
                if cand in s:
                    return c
        return None

    for k in range(3):
        try:
            r = SESSION.get(url, headers=headers, timeout=25, verify=verify_ssl)
            ctype = (r.headers.get("Content-Type") or "").lower()
            if r.status_code != 200:
                log(f"TWSE OpenAPI v1 non-200: {r.status_code}, ct={ctype}, head={r.text[:80]!r}")
                continue

            try:
                j = r.json()
            except Exception as e:
                log(f"TWSE OpenAPI v1 json decode failed: {repr(e)} ct={ctype} head={r.text[:120]!r}")
                continue

            if not isinstance(j, list) or len(j) == 0:
                log(f"TWSE OpenAPI v1 empty/non-list payload: type={type(j)}")
                continue

            df = pd.DataFrame(j)

            col_code = _find_col(df, ["股票代號", "證券代號", "證券代碼"])
            c_m_today = _find_col(df, ["融資今日餘額", "融資餘額"])
            c_s_today = _find_col(df, ["融券今日餘額", "融券餘額"])

            c_s_prev = _find_col(df, ["融券前日餘額"])
            c_s_buy = _find_col(df, ["融券買進"])
            c_s_sell = _find_col(df, ["融券賣出"])
            c_s_repay = _find_col(df, ["融券現券償還", "融券償還"])

            if not (col_code and c_m_today):
                log(f"TWSE OpenAPI v1 missing base cols: cols={list(df.columns)[:12]}")
                continue

            sym = df[col_code].astype(str).str.strip()
            margin_today = pd.to_numeric(df[c_m_today].astype(str).str.replace(",", ""), errors="coerce")

            if c_s_today:
                short_today = pd.to_numeric(df[c_s_today].astype(str).str.replace(",", ""), errors="coerce")
            else:
                if not (c_s_prev and c_s_buy and c_s_sell):
                    log(f"TWSE OpenAPI v1 missing short cols for derive: cols={list(df.columns)[:12]}")
                    continue
                s_prev = pd.to_numeric(df[c_s_prev].astype(str).str.replace(",", ""), errors="coerce")
                s_buy = pd.to_numeric(df[c_s_buy].astype(str).str.replace(",", ""), errors="coerce")
                s_sell = pd.to_numeric(df[c_s_sell].astype(str).str.replace(",", ""), errors="coerce")
                if c_s_repay:
                    s_repay = pd.to_numeric(df[c_s_repay].astype(str).str.replace(",", ""), errors="coerce")
                else:
                    s_repay = 0.0
                short_today = s_prev + s_sell - s_buy - s_repay

            out = pd.DataFrame({
                "symbol": sym,
                "margin_balance": margin_today,
                "short_balance": short_today,
                "market": "TWSE",
            })
            out = out[out["symbol"].str.fullmatch(r"\d{4}")].copy()
            out = out.dropna(subset=["margin_balance", "short_balance"])
            if out.empty:
                log("TWSE OpenAPI v1 parsed empty after filter/na")
                continue
            return out

        except Exception as e:
            log(f"TWSE OpenAPI v1 request failed: {repr(e)} (try={k+1})")
            continue

    return pd.DataFrame()


def fetch_margin_short_ratio_map(signal_date: _dt.date, lookback_days: int = 30) -> dict[tuple[str, str], float]:
    """
    券資比(%) = 融券餘額 / 融資餘額 * 100

    來源策略：
    - 上市：TWSE OpenAPI v1（最新一筆，含推導融券今日餘額）
    - 上櫃：TPEx JSON（回溯 lookback_days 找到最近有資料日）

    同步寫入全域 MARGIN_RATIO_META：OK / DIV0 / NA / NO_DATA
    """
    global MARGIN_RATIO_META
    MARGIN_RATIO_META = {}

    ratio: dict[tuple[str, str], float] = {}

    # ---- TWSE (latest) ----
    try:
        tw = _twse_openapi_mi_margn_df()
        if tw is not None and not tw.empty:
            for r in tw.itertuples(index=False):
                key = (str(r.symbol).strip(), str(r.market).strip())
                if pd.isna(r.margin_balance) or pd.isna(r.short_balance):
                    MARGIN_RATIO_META[key] = "NA"

            div0 = tw[(tw["margin_balance"].fillna(0) == 0) & (tw["short_balance"].notna())]
            for r in div0.itertuples(index=False):
                MARGIN_RATIO_META[(str(r.symbol).strip(), str(r.market).strip())] = "DIV0"

            tw2 = tw[(tw["margin_balance"].notna()) & (tw["short_balance"].notna()) & (tw["margin_balance"] != 0)].copy()
            tw2["ratio_pct"] = (tw2["short_balance"] / tw2["margin_balance"]) * 100.0
            tw2 = tw2.dropna(subset=["ratio_pct"])
            for r in tw2.itertuples(index=False):
                key = (str(r.symbol).strip(), str(r.market).strip())
                ratio[key] = float(r.ratio_pct)
                MARGIN_RATIO_META[key] = "OK"
            log(f"券資比(TWSE OpenAPI v1) loaded n={sum(1 for k in ratio.keys() if isinstance(k, tuple) and len(k)==2 and k[1]=='TWSE')}")
        else:
            log("TWSE OpenAPI v1 empty payload")
    except Exception as e:
        log(f"TWSE ratio load failed: {repr(e)}")

    # ---- TPEx (backtrack) ----
    # ---- TPEx latest HTML (no-date) ----
    try:
        two_latest = _tpex_margin_latest_html_df()
        if two_latest is not None and not two_latest.empty:
            for r in two_latest.itertuples(index=False):
                key = (str(r.symbol).strip(), str(r.market).strip())
                if pd.isna(r.margin_balance) or pd.isna(r.short_balance):
                    MARGIN_RATIO_META[key] = "NA"
            div0 = two_latest[(two_latest["margin_balance"].fillna(0) == 0) & (two_latest["short_balance"].notna())]
            for r in div0.itertuples(index=False):
                MARGIN_RATIO_META[(str(r.symbol).strip(), str(r.market).strip())] = "DIV0"
            two2 = two_latest[(two_latest["margin_balance"].notna()) & (two_latest["short_balance"].notna()) & (two_latest["margin_balance"] != 0)].copy()
            two2["ratio_pct"] = (two2["short_balance"] / two2["margin_balance"]) * 100.0
            two2 = two2.dropna(subset=["ratio_pct"])
            for r in two2.itertuples(index=False):
                key = (str(r.symbol).strip(), str(r.market).strip())
                ratio[key] = float(r.ratio_pct)
                MARGIN_RATIO_META[key] = "OK"
            n_two = sum(1 for k in ratio.keys() if isinstance(k, tuple) and len(k)==2 and k[1]=='TWO')
            log(f"券資比(TWO latest HTML) loaded n={n_two}")
            if n_two > 0:
                return ratio
            else:
                log("TWO latest HTML produced 0 valid ratios; fallback to backtrack")
    except Exception as e:
        log(f"TPEx latest HTML ratio load failed: {repr(e)}")

    used_date = None
    two_df = pd.DataFrame()
    for back in range(1, lookback_days + 1):
        d = signal_date - _dt.timedelta(days=back)
        df = _tpex_margin_df(f"{d.year - 1911}/{d.month:02d}/{d.day:02d}")
        if df is not None and not df.empty:
            used_date = d
            two_df = df
            break

    if used_date is not None and two_df is not None and not two_df.empty:
        for r in two_df.itertuples(index=False):
            key = (str(r.symbol).strip(), str(r.market).strip())
            if pd.isna(r.margin_balance) or pd.isna(r.short_balance):
                MARGIN_RATIO_META[key] = "NA"

        div0 = two_df[(two_df["margin_balance"].fillna(0) == 0) & (two_df["short_balance"].notna())]
        for r in div0.itertuples(index=False):
            MARGIN_RATIO_META[(str(r.symbol).strip(), str(r.market).strip())] = "DIV0"

        two2 = two_df[(two_df["margin_balance"].notna()) & (two_df["short_balance"].notna()) & (two_df["margin_balance"] != 0)].copy()
        two2["ratio_pct"] = (two2["short_balance"] / two2["margin_balance"]) * 100.0
        two2 = two2.dropna(subset=["ratio_pct"])
        for r in two2.itertuples(index=False):
            key = (str(r.symbol).strip(), str(r.market).strip())
            ratio[key] = float(r.ratio_pct)
            MARGIN_RATIO_META[key] = "OK"
        log(f"券資比(TWO) loaded date={used_date.isoformat()} n={sum(1 for k in ratio.keys() if isinstance(k, tuple) and len(k)==2 and k[1]=='TWO')}")
    else:
        log(f"券資比(TWO) not available in last {lookback_days} days.")

    return ratio
# =======================================

def prefilter_by_liquidity(tickers: list[str]) -> list[str]:
    """
    v6.3: 兩階段快篩 - Stage1 只下載近 STAGE1_PERIOD，用成交量/資料可用性篩掉大部分股票
    回傳：進入 Stage2 的 tickers
    """
    if not ENABLE_TWO_STAGE_SCREEN:
        return tickers
    log(f"Stage1 prefilter: period={STAGE1_PERIOD}, universe={len(tickers)}")
    vol_map = {}  # ticker -> avg_volume
    for bi, batch in enumerate(batched(tickers, BATCH_SIZE), start=1):
        log(f"Stage1 batch {bi}: {len(batch)} tickers")
        try:
            data = yf_download_with_retry(batch, STAGE1_PERIOD)
        except Exception as e:
            log(f"Stage1 batch download failed: {repr(e)}")
            continue

        for t in batch:
            try:
                if isinstance(data.columns, pd.MultiIndex):
                    df = data[t].dropna(how="all")
                else:
                    df = data.dropna(how="all")
                if df is None or df.empty or "Volume" not in df.columns:
                    INVALID_TICKERS.add(t)
                    continue
                av = float(df["Volume"].tail(5).mean())
                if av > 0:
                    vol_map[t] = av
            except Exception:
                continue
        time.sleep(SLEEP_BETWEEN_YF_BATCH)

    if not vol_map:
        log("Stage1 prefilter got no data; fallback to full universe.")
        return tickers

    # keep above MIN_AVG_VOLUME then top N
    items = [(t, v) for t, v in vol_map.items() if v >= MIN_AVG_VOLUME]
    items.sort(key=lambda x: x[1], reverse=True)
    sel = [t for t, _ in items[:TOPN_LIQUID]]
    log(f"Stage1 selected for Stage2: {len(sel)} tickers (min_vol={MIN_AVG_VOLUME}, topN={TOPN_LIQUID})")
    return sel if sel else tickers

def download_histories(tickers: list[str], period: str = "6mo") -> dict[str, pd.DataFrame]:
    out, missing = {}, []
    for t in tickers:
        smr = None  # v6.3.18.4: default init
        risk_note_extra = ""
        c = load_cached_history(t)
        if c is not None and not c.empty:
            out[t] = c
        else:
            missing.append(t)
    for bi, batch in enumerate(batched(missing, BATCH_SIZE), start=1):
        log(f"Downloading batch {bi}: {len(batch)} tickers")
        if not batch:
            continue
        try:
            data = yf_download_with_retry(batch, period)
        except Exception:
            continue
        for t in batch:
            try:
                df = data[t].dropna(how="all") if isinstance(data.columns, pd.MultiIndex) else data.dropna(how="all")
                if df is not None and not df.empty:
                    out[t] = df
                    save_cached_history(t, df)
                else:
                    INVALID_TICKERS.add(t)
            except Exception:
                continue
        time.sleep(SLEEP_BETWEEN_YF_BATCH)
    return out


def compute_indicators(hist: pd.DataFrame):
    if hist is None or hist.empty or len(hist) < 60:
        return None
    if float(hist["Volume"].tail(20).mean()) < MIN_AVG_VOLUME:
        return None
    close, high, low, vol = hist["Close"], hist["High"], hist["Low"], hist["Volume"]
    ma20 = close.rolling(20).mean()
    bias20 = (close - ma20) / ma20 * 100
    prev_close = close.shift(1)
    tr = pd.concat([(high - low), (high - prev_close).abs(), (low - prev_close).abs()], axis=1).max(axis=1)
    atr20 = tr.rolling(20).mean()
    range_20 = high.rolling(20).max() - low.rolling(20).min()
    range_60 = high.rolling(60).max() - low.rolling(60).min()
    volatility_ratio = range_20 / range_60
    volume_ratio = vol.rolling(5).mean() / vol.rolling(20).mean()
    out = {
        "close": float(close.iloc[-1]),
        "ma20": float(ma20.iloc[-1]),
        "bias20": float(bias20.iloc[-1]),
        "support_1m": float(low.tail(20).min()),
        "atr20": float(atr20.iloc[-1]),
        "volatility_ratio": float(volatility_ratio.iloc[-1]),
        "volume_ratio": float(volume_ratio.iloc[-1]),
    }
    if math.isnan(out["atr20"]) or out["atr20"] <= 0 or math.isnan(out["bias20"]) or math.isnan(out["ma20"]):
        return None
    return out


def calc_market_regime(index_hist: pd.DataFrame) -> str:
    if index_hist is None or index_hist.empty or len(index_hist) < 60:
        return "RANGE"
    close, high, low = index_hist["Close"], index_hist["High"], index_hist["Low"]
    ma60 = close.rolling(60).mean()
    ma20 = close.rolling(20).mean()
    bias20 = (close - ma20) / ma20 * 100
    range_20 = high.rolling(20).max() - low.rolling(20).min()
    range_60 = high.rolling(60).max() - low.rolling(60).min()
    trend = abs(float(close.iloc[-1]) - float(ma60.iloc[-1])) / float(ma60.iloc[-1])
    vol_ratio = float(range_20.iloc[-1]) / float(range_60.iloc[-1])
    b = float(bias20.iloc[-1])
    if vol_ratio >= 0.9 and abs(b) > 6:
        return "STRESS"
    if trend >= 0.03 and vol_ratio >= 0.7:
        return "TREND"
    return "RANGE"


def calc_otc_short_pressure(volatility_ratio: float, volume_ratio: float) -> float:
    return volatility_ratio * 0.6 + volume_ratio * 0.4

def _squeeze_tw_guard(ind: dict, short_margin_ratio: float | None) -> bool:
    return (
        short_margin_ratio is not None
        and short_margin_ratio >= TH_SQUEEZE_TW
        and ind["volatility_ratio"] >= TH_SQUEEZE_VOL_RATIO
        and ind["volume_ratio"] >= TH_SQUEEZE_VOLRATIO_VOL
        and ind["close"] >= ind["ma20"]
        and ind["bias20"] >= TH_SQUEEZE_MIN_BIAS
    )

def tag_strategy_complete(market: str, ind: dict, short_margin_ratio: float | None, otc_short_pressure: float):
    if market == "TW" and _squeeze_tw_guard(ind, short_margin_ratio):
        return "SQUEEZE_TW"
    if market == "TWO" and otc_short_pressure >= TH_SQUEEZE_OTC_PROXY:
        return "SQUEEZE_OTC"
    if ind["close"] < ind["support_1m"] * (1.0 + TH_SUPPORT_TOL):
        return None
    if ind["bias20"] <= TH_BIAS_DEEP:
        return "HIGH_MARGIN_MEAN_REVERT"
    if ind["bias20"] <= TH_BIAS_MEAN_REVERT:
        return "MEAN_REVERT"
    return None


def _load_prev_weights():
    if not os.path.exists(WEIGHT_STATE_FILE):
        return None
    try:
        with open(WEIGHT_STATE_FILE, "r", encoding="utf-8") as f:
            return (json.load(f) or {}).get("weights", None)
    except Exception:
        return None

def _save_prev_weights(w):
    try:
        with open(WEIGHT_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump({"updated": _today_str(), "weights": w}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _cap_and_renorm(w):
    w2 = {k: min(float(v), MAX_STRATEGY_WEIGHT) for k, v in w.items()}
    s = sum(w2.values())
    return {k: v / s for k, v in w2.items()} if s > 0 else w2

def load_weight_inputs_from_summary(path: str):
    if not os.path.exists(path):
        return None
    try:
        df = pd.read_excel(path, sheet_name="ByStrategy")
    except Exception:
        return None
    if "strategy" not in [c.lower() for c in df.columns]:
        return None
    strategy_col = next(c for c in df.columns if c.lower() == "strategy")
    ann_col = next((c for c in df.columns if c.lower().startswith("annualized")), None)
    mdd_col = next((c for c in df.columns if c.lower().startswith("mdd")), None)
    trades_col = next((c for c in df.columns if c.lower().startswith("trades")), None)
    if ann_col is None or mdd_col is None:
        return None
    cols = [strategy_col, ann_col, mdd_col] + ([trades_col] if trades_col else [])
    df = df[cols].copy()
    df.columns = ["strategy", "annualized_pct", "mdd_pct"] + (["trades"] if trades_col else [])
    df["annualized_pct"] = pd.to_numeric(df["annualized_pct"], errors="coerce")
    df["mdd_pct"] = pd.to_numeric(df["mdd_pct"], errors="coerce")
    df["trades"] = pd.to_numeric(df["trades"], errors="coerce") if "trades" in df.columns else pd.NA
    df = df.dropna(subset=["strategy", "annualized_pct", "mdd_pct"])
    if df.empty:
        return None
    df["score_raw"] = (df["annualized_pct"].clip(lower=0) / df["mdd_pct"].abs().clip(lower=1.0))
    df["score_raw"] = df["score_raw"].replace([math.inf, -math.inf], pd.NA).fillna(0.0)
    mask = df["trades"].notna() & (df["trades"] < MIN_TRADES_FOR_WEIGHT)
    df.loc[mask, "score_raw"] *= (df.loc[mask, "trades"] / MIN_TRADES_FOR_WEIGHT).clip(lower=0, upper=1)
    out = {}
    for _, r in df.iterrows():
        out[str(r["strategy"])] = {
            "annualized_pct": float(r["annualized_pct"]),
            "mdd_pct": float(r["mdd_pct"]),
            "trades": None if pd.isna(r["trades"]) else float(r["trades"]),
            "score_raw": float(r["score_raw"]),
        }
    return out

def compute_weights_with_trace(strategies_today: list[str]):
    inputs = load_weight_inputs_from_summary(PERF_SUMMARY_FILE)
    if inputs is None:
        # v6.3.16: 若無 performance_summary（檔案不存在或無 CLOSED trades），仍回填欄位避免整欄空白
        # 嘗試自動產出 performance_summary.xlsx（若有 CLOSED trades）
        try:
            if not os.path.exists(PERF_SUMMARY_FILE):
                import performance_summary as _ps
                _ps.main()
        except Exception:
            pass
        # 再讀一次
        inputs2 = load_weight_inputs_from_summary(PERF_SUMMARY_FILE)
        if inputs2 is not None:
            inputs = inputs2
        else:
            prev = _load_prev_weights() or {}
            w = 1.0 / len(strategies_today) if strategies_today else 1.0
            trace = {}
            for s in strategies_today:
                trace[s] = {
                    "weight_mode": "EQUAL",
                    "weight_source_file": f"{PERF_SUMMARY_FILE} (missing/no CLOSED trades)",
    "weight_score_raw": (safe_float(safe_get(locals(), "weight_score_raw", 0.0, SAFE_MODE), 0.0, SAFE_MODE) + safe_float(safe_get(locals(), "score_penalty", 0.0, SAFE_MODE), 0.0, SAFE_MODE)),
                    "weight_raw": w,
                    "weight_prev": prev.get(s),
                    "weight_used": w,
                    "weight_alpha": WEIGHT_SMOOTH_ALPHA,
                    "min_trades": MIN_TRADES_FOR_WEIGHT,
                    "trades_used": 0,
                    "annualized_pct": 0.0,
                    "mdd_pct": 0.0,
                }
            return ({s: w for s in strategies_today}, trace, "EQUAL")
    score = {s: float(inputs[s]["score_raw"]) for s in inputs.keys()}
    total = sum(score.values())
    if total <= 0:
        w = 1.0 / len(strategies_today) if strategies_today else 1.0
        return ({s: w for s in strategies_today}, {s: {"weight_mode":"EQUAL"} for s in strategies_today}, "EQUAL")

    raw_all = _cap_and_renorm({k: v / total for k, v in score.items()})
    prev = _load_prev_weights()
    mode = "DYNAMIC_SMOOTHED" if (ENABLE_WEIGHT_SMOOTHING and prev is not None) else "DYNAMIC"

    # v6.3.15: display meta for weights
    weight_source_file_display = str(weight_source_file) if "weight_source_file" in locals() else ""
    prev_w = prev_weights if "prev_weights" in locals() else {}

    if ENABLE_WEIGHT_SMOOTHING and prev is not None:
        alpha = WEIGHT_SMOOTH_ALPHA
        keys = set(raw_all.keys()) | set(prev.keys())
        sm = {k: alpha*float(raw_all.get(k,0.0)) + (1-alpha)*float(prev.get(k,0.0)) for k in keys}
        sm = _cap_and_renorm({k:v for k,v in sm.items() if v>0})
    else:
        sm = raw_all

    eps = 1e-6
    used = {s: float(sm.get(s, eps)) for s in strategies_today}
    ssum = sum(used.values())
    used = {k: v/ssum for k,v in used.items()} if ssum>0 else {s: 1.0/len(strategies_today) for s in strategies_today}

    _save_prev_weights(used)

    trace = {}
    for s in strategies_today:
        inp = inputs.get(s, {})
        trace[s] = {
            "weight_mode": mode,
            "weight_source_file": PERF_SUMMARY_FILE,
    "weight_score_raw": (safe_float(safe_get(locals(), "weight_score_raw", 0.0, SAFE_MODE), 0.0, SAFE_MODE) + safe_float(safe_get(locals(), "score_penalty", 0.0, SAFE_MODE), 0.0, SAFE_MODE)),
            "weight_raw": raw_all.get(s),
            "weight_prev": None if prev is None else prev.get(s),
            "weight_used": used.get(s),
            "weight_alpha": WEIGHT_SMOOTH_ALPHA,
            "min_trades": MIN_TRADES_FOR_WEIGHT,
            "trades_used": inp.get("trades"),
            "annualized_pct": inp.get("annualized_pct"),
            "mdd_pct": inp.get("mdd_pct"),
        }
    return used, trace, mode


def calc_position_size(total_capital: float, strategy_weight: float, atr20: float, close: float, market_regime: str) -> float:
    vol_pct = atr20 / close
    if vol_pct <= 0:
        return 0.0
    account_risk = total_capital * MAX_ACCOUNT_RISK
    strategy_risk = account_risk * strategy_weight * REGIME_RISK_MULTIPLIER.get(market_regime, 1.0)
    max_trade_risk = strategy_risk * PER_TRADE_RISK_RATIO
    return float(min(max_trade_risk / vol_pct, total_capital * 0.5))



def format_excel_sheet(file_path: str, hide_headers: list[str] | None = None) -> None:
    """
    v6.3.7：格式化 Excel（凍結首列/篩選/欄寬/醒目風險提醒/隱藏指定欄位）
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(file_path)
        ws = wb.active

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto filter
        ws.auto_filter.ref = ws.dimensions

        # Header style
        header_fill = PatternFill("solid", fgColor="EDEDED")
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Hide specified columns by header text (e.g., Yahoo代碼)
        if hide_headers:
            header_to_col = {str(cell.value).strip(): j for j, cell in enumerate(ws[1], start=1)}
            for h in hide_headers:
                if h in header_to_col:
                    col_letter = get_column_letter(header_to_col[h])
                    ws.column_dimensions[col_letter].hidden = True

        # Column widths (simple heuristic)
        for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
            max_len = 0
            for cell in col[: min(ws.max_row, 200)]:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            width = min(max(10, max_len + 2), 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Risk highlighting
        risk_col = None
        for j, cell in enumerate(ws[1], start=1):
            if str(cell.value).strip() == "風險提醒":
                risk_col = j
                break
        if risk_col:
            fill_risk = PatternFill("solid", fgColor="FFF2CC")  # light yellow
            fill_high = PatternFill("solid", fgColor="F8CBAD")  # light red
            for r in range(2, ws.max_row + 1):
                v = ws.cell(r, risk_col).value
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = fill_risk
                if "HIGH" in s.upper() or "高" in s:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = fill_high

        wb.save(file_path)
    except Exception as e:
        log("Excel format skipped: " + repr(e))

def send_email_with_attachment(to_email: str, file_path: str, subject: str, body: str) -> None:
    app_pw = os.environ.get(APP_PASSWORD_ENV, "").strip()
    if not app_pw:
        raise RuntimeError(f"Missing Gmail App Password env var: {APP_PASSWORD_ENV}")
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = SENDER_EMAIL
    msg["To"] = to_email
    msg.set_content(body)
    with open(file_path, "rb") as f:
        data = f.read()
    msg.add_attachment(data, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=os.path.basename(file_path))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(SENDER_EMAIL, app_pw)
        smtp.send_message(msg)




def compute_composite_score(df: pd.DataFrame):
    """Compute composite score for ranking (daily + top20 consistent).
    Adds: traded value rank and vol*liquidity interaction penalty.
    """
    def _rank01(series, higher_better=True):
        s = pd.to_numeric(series, errors="coerce")
        # if scalar (float/int), expand to constant series
        if not hasattr(s, "isna"):
            s = pd.Series([float(s) if s is not None else float("nan")] * len(df), index=df.index)
        if s.isna().all():
            return pd.Series([0.0]*len(df), index=df.index)
        r = s.rank(pct=True)
        if not higher_better:
            r = 1.0 - r
        return r.fillna(0.0)

    pos_col = next((c for c in ["建議部位","position","target_position","final_position","target_weight","position_weight"] if c in df.columns), None)
    score_col = next((c for c in ["策略分數","strategy_score","score"] if c in df.columns), None)
    bias_col = next((c for c in ["乖離率(%)","bias(%)","BIAS(%)","bias"] if c in df.columns), None)
    turn_col = next((c for c in ["周轉率(%)","turnover_rate(%)"] if c in df.columns), None)
    tv_col = next((c for c in ["成交值(元)","traded_value_ntd"] if c in df.columns), None)
    vol_col = next((c for c in ["vol_annual","年化波動"] if c in df.columns), None)
    price_col = next((c for c in ["進場價","entry_price","close","Close","收盤價"] if c in df.columns), None)

    w_pos = _rank01(df[pos_col] if pos_col else 0.0, True)
    w_score = _rank01(df[score_col] if score_col else 0.0, True)

    if bias_col:
        b = pd.to_numeric(df[bias_col], errors="coerce").fillna(0.0).abs()
        w_bias = _rank01(b, True)
    else:
        w_bias = pd.Series([0.0]*len(df), index=df.index)

    if turn_col:
        tr = pd.to_numeric(df[turn_col], errors="coerce").fillna(0.0)
        w_turn = _rank01(tr, True)
    else:
        w_turn = pd.Series([0.0]*len(df), index=df.index)

    if tv_col:
        tv = pd.to_numeric(df[tv_col], errors="coerce").fillna(0.0)
        w_tv = _rank01(tv, True)
    else:
        w_tv = pd.Series([0.0]*len(df), index=df.index)

    prices = pd.to_numeric(df[price_col], errors="coerce") if price_col else pd.Series([float("nan")]*len(df), index=df.index)
    tvv = pd.to_numeric(df[tv_col], errors="coerce") if tv_col else pd.Series([float("nan")]*len(df), index=df.index)
    vola = pd.to_numeric(df[vol_col], errors="coerce") if vol_col else pd.Series([float("nan")]*len(df), index=df.index)

    pen=[]
    for pr, tv, va in zip(prices.tolist(), tvv.tolist(), vola.tolist()):
        p=0.0
        try:
            thr_amt = amount_threshold_by_price(pr)
            if (tv is not None) and (not (isinstance(tv,float) and tv!=tv)):
                if float(tv) < float(thr_amt):
                    p += 0.30
            if (va is not None) and (not (isinstance(va,float) and va!=va)) and (tv is not None) and (not (isinstance(tv,float) and tv!=tv)):
                if float(va) > 0.50 and float(tv) < float(thr_amt)*1.5:
                    p += 0.30
        except Exception:
            pass
        pen.append(p)
    penalty = pd.Series(pen, index=df.index)

    comp = 0.25*w_pos + 0.25*w_score + 0.20*w_bias + 0.10*w_turn + 0.20*w_tv - penalty
    out = df.copy()
    out["綜合分數"] = comp.round(4)
    out["流動性扣分"] = penalty.round(2)
    out["成交值排名"] = w_tv.round(4)
    return out



def _extract_stock_code_any(x: str):
    try:
        s = str(x).strip()
        m = re.search(r"(\d{4,6})", s)
        return m.group(1) if m else None
    except Exception:
        return None


def compute_turnover_rate_percent(df: pd.DataFrame, shares_map: dict):
    """Compute turnover_rate(%) = volume / shares_outstanding * 100."""
    out = df.copy()
    vcol = next((c for c in ["volume","Volume","成交量"] if c in out.columns), None)
    tcol = next((c for c in ["ticker","symbol","Yahoo代碼","yahoo_code","yahoo_ticker"] if c in out.columns), None)
    if vcol is None or tcol is None or not shares_map:
        out["turnover_rate(%)"] = float("nan")
        return out
    vols = pd.to_numeric(out[vcol], errors="coerce")
    codes = out[tcol].apply(_extract_stock_code_any)
    sh = codes.map(lambda c: shares_map.get(str(c)) if c else None)
    shs = pd.to_numeric(pd.Series(sh, index=out.index), errors="coerce")
    out["turnover_rate(%)"] = (vols / shs * 100.0)
    return out



def fetch_official_daily_volume_map(trade_date):
    """Return dict: stock_code(str digits) -> volume(int shares) using official TWSE/TPEx open APIs."""
    vol_map = {}
    # TWSE OpenAPI: daily close quotes (all) includes 成交股數
    try:
        url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
        r = SESSION.get(url, timeout=30)
        if r.status_code == 200 and r.text.strip().startswith("["):
            data = r.json()
            for row in data:
                code = str(row.get("證券代號") or row.get("股票代號") or "").strip()
                if not code.isdigit():
                    continue
                v = row.get("成交股數") or row.get("成交量") or row.get("成交股數(股)")
                try:
                    vol = int(str(v).replace(",", "").strip())
                    if vol > 0:
                        vol_map[code] = vol
                except Exception:
                    continue
    except Exception as e:
        log(f"official volume map fetch (TWSE openapi) failed: {repr(e)}")

    # TPEx OpenAPI: daily close quotes (mainboard)
    try:
        roc_year = trade_date.year - 1911
        roc_date = f"{roc_year}/{trade_date.month:02d}/{trade_date.day:02d}"
        url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_daily_close_quotes"
        params = {"l": "zh-tw", "d": roc_date, "s": "0,asc,0"}
        r = SESSION.get(url, params=params, timeout=30)
        if r.status_code == 200 and r.text.strip().startswith("["):
            data = r.json()
            for row in data:
                code = str(row.get("代號") or row.get("股票代號") or row.get("證券代號") or "").strip()
                if not code.isdigit():
                    continue
                v = row.get("成交股數") or row.get("成交量") or row.get("成交股數(股)")
                try:
                    vol = int(str(v).replace(",", "").strip())
                    if vol > 0:
                        vol_map[code] = vol
                except Exception:
                    continue
    except Exception as e:
        log(f"official volume map fetch (TPEx) failed: {repr(e)}")

    return vol_map



# --- Yahoo volume fallback (candidate-only) ---
try:
    import yfinance as yf
except Exception:
    yf = None


def fetch_yahoo_volume_map(tickers, trade_date, chunk_size=80, pause_sec=1.0):
    """Fetch volume for tickers from Yahoo (yfinance). Returns dict: ticker -> volume(int).
    Only for a limited candidate list to reduce rate limits.
    """
    if yf is None:
        return {}

    tset = [t for t in list(dict.fromkeys(tickers)) if isinstance(t, str) and t.strip()]
    if not tset:
        return {}

    vol_map = {}
    start = (trade_date - timedelta(days=10)).strftime("%Y-%m-%d")
    end = (trade_date + timedelta(days=1)).strftime("%Y-%m-%d")

    for i in range(0, len(tset), chunk_size):
        chunk = tset[i:i+chunk_size]
        try:
            df = yf.download(
                tickers=" ".join(chunk),
                start=start,
                end=end,
                group_by="ticker",
                auto_adjust=False,
                threads=False,
                progress=False,
            )
            if df is None or len(df) == 0:
                time.sleep(pause_sec)
                continue

            idx_dates = [d.date() for d in df.index]

            def _pick_last(series):
                best = None
                for d, v in zip(idx_dates, series.values):
                    try:
                        if d <= trade_date and v is not None and not pd.isna(v):
                            best = v
                    except Exception:
                        continue
                return best

            if isinstance(df.columns, pd.MultiIndex):
                for t in chunk:
                    try:
                        if (t, "Volume") in df.columns:
                            v = _pick_last(df[(t, "Volume")])
                        else:
                            v = None
                        if v is not None and float(v) > 0:
                            vol_map[t] = int(float(v))
                    except Exception:
                        continue
            else:
                if "Volume" in df.columns and len(chunk) == 1:
                    v = _pick_last(df["Volume"])
                    if v is not None and float(v) > 0:
                        vol_map[chunk[0]] = int(float(v))
        except Exception as e:
            log(f"yahoo volume fetch failed chunk {i//chunk_size+1}: {repr(e)}")

        time.sleep(pause_sec)

    return vol_map

def main():
    global decision_path, full_path, today, market_regime
    df_view_dec = pd.DataFrame()  # init to avoid UnboundLocalError
    now = _dt.datetime.now()
    today_date = now.date()
    today = _dt.date.today()
    log('=== daily_auto_run_final start ===')
    log(f'cwd={os.getcwd()}')
    os.makedirs(LOCAL_EXCEL_FOLDER, exist_ok=True)
    os.makedirs(CACHE_DIR, exist_ok=True)

    uni = pd.concat([fetch_listed_stocks(), fetch_otc_stocks()], ignore_index=True)

    # v6.3.1: 清理股票池（只保留 4 碼普通股）
    uni["symbol"] = uni["symbol"].astype(str).str.strip()
    uni = uni[uni["symbol"].str.fullmatch(r"\d{4}")].copy()
    log(f"Universe after clean (4-digit only): {len(uni)}")

    # v6.3.2: 由 ISIN 名單補上中文名稱（不額外打 Yahoo）
    try:
        isin_tw = _fetch_isin_universe(2)
        isin_two = _fetch_isin_universe(4)
        isin_all = pd.concat([isin_tw, isin_two], ignore_index=True)
        set_name_map(isin_all)
        if 'name_zh' not in isin_all.columns:
            isin_all['name_zh'] = ''
        uni = uni.merge(isin_all[[c for c in ['symbol','market','name_zh'] if c in isin_all.columns]], on=['symbol','market'], how='left')
        uni['name_zh'] = uni['name_zh'].fillna('')
        log('Name enrichment done.')
        # ensure name_zh exists (v6.3.8)
        if 'name_zh' not in uni.columns:
            uni['name_zh'] = ''
        uni['name_zh'] = uni['name_zh'].fillna('')
    except Exception as e:
        uni['name_zh'] = ''
        log('Name enrichment skipped: ' + repr(e))
        # v6.3.20: load ratio_map once per run (券資比 T+1 backtrack)
        try:
            ratio_map = fetch_margin_short_ratio_map(today, lookback_days=30)
            log(f"ratio_map size (T+1): {len(ratio_map)}")
        except Exception as _e:
            ratio_map = {}
            log("ratio_map load failed: " + repr(_e))


    # v6.3.11: post-enrichment safeguard (即使前段 try/except 失敗也強制補上 name_zh)
    if "name_zh" not in uni.columns:
        try:
            _tw = _fetch_isin_universe(2)
            _two = _fetch_isin_universe(4)
            _all = pd.concat([_tw, _two], ignore_index=True)
            set_name_map(_all)
            if "name_zh" not in _all.columns:
                _all["name_zh"] = ""
            uni = uni.merge(_all[["symbol","market","name_zh"]], on=["symbol","market"], how="left")
            uni["name_zh"] = uni["name_zh"].fillna("")
            log("Name enrichment (safeguard) done.")
        except Exception as _e:
            uni["name_zh"] = ""
            log("Name enrichment (safeguard) failed: " + repr(_e))

    # v6.3.20.11: legacy TWSE CSV ratio_map overwrite removed (keeps tuple-key ratio_map)
    log("legacy TWSE CSV ratio_map overwrite disabled")

    log(f"ratio_map size (TWSE keys): {sum(1 for k in ratio_map.keys() if isinstance(k, tuple) and len(k)==2 and k[1]=='TWSE')}")
    idx_hist = yf.Ticker(INDEX_TICKER).history(period="6mo")
    market_regime = calc_market_regime(idx_hist)

    tickers, meta = [], {}
    for _, r in uni.iterrows():
        sym, market = str(r["symbol"]), r["market"]
        t = f"{sym}.TW" if market == "TW" else f"{sym}.TWO"
        tickers.append(t)
        meta[t] = (sym, market, str(r.get("name_zh","")))

    tickers2 = prefilter_by_liquidity(tickers)


    latest_volume_map = {}  # yahoo_symbol -> latest volume
    histories = download_histories(tickers2, period=STAGE2_PERIOD)

    rows = []
    for t, hist in histories.items():
        sym, market, name_zh = meta.get(t, ("", "", ""))
        if not name_zh:
            name_zh = NAME_MAP.get((str(sym).strip(), str(market).strip()), "")
        if not sym:
            continue
        ind = compute_indicators(hist)
        if ind is None:
            continue
        smr = ratio_map.get((str(sym).strip(), normalize_market_code(market)), None)
        # v6.3.21.0: SMR status detail
        smr_key = (str(sym).strip(), normalize_market_code(market))
        smr_meta = MARGIN_RATIO_META.get(smr_key, "NO_DATA")
        smr_label = smr_text_label(smr, smr_meta)
        smr_light = smr_traffic_light(smr_label)
        otc_p = calc_otc_short_pressure(ind["volatility_ratio"], ind["volume_ratio"]) if market == "TWO" else 0.0
        strategy = tag_strategy_complete(market, ind, smr, otc_p)
        if strategy is None:
            continue
        # v6.3.18.4: 進場前風險規則 - 券資比缺值 => 禁止下單（部位=0，風險提醒加註）
        if smr is None or smr == "" or (isinstance(smr, float) and pd.isna(smr)):
            risk_note_extra = "⚠券資比缺值：禁止下單"
            try:
                position_size = 0
            except Exception:
                pass

        score_penalty = 0.0
        # v6.3.19 SAFE: 券資比規則（缺值或>=50%：不剔除，但部位=0+禁止下單；>=30% 扣分）
        smr = None
        smr_status = smr_meta
        try:
            smr = ratio_map.get((str(sym).strip(), normalize_market_code(market)), None)
        except Exception:
            smr = None
        smr_val = safe_float(smr, default=None, safe_mode=True) if (smr is not None and smr != "") else None
        if smr_val is None:
            smr_status = smr_meta
            risk_note_extra = (risk_note_extra + " | " if risk_note_extra else "") + "⚠券資比缺值/異常：禁止下單"
            try:
                position_size = 0
            except Exception:
                pass
        else:
            if smr_val >= 50:
                smr_status = "BLOCK_50"
                risk_note_extra = (risk_note_extra + " | " if risk_note_extra else "") + f"⚠券資比極高({smr_val:.1f}%) 禁止下單"
                try:
                    position_size = 0
                except Exception:
                    pass
            if smr_val >= 30:
                if not smr_status:
                    smr_status = "HIGH_30"
                score_penalty += -1.0
                risk_note_extra = ((locals().get("risk_note_extra","") + " | ") if str(locals().get("risk_note_extra","")).strip() else "") + f"⚠券資比高({smr_val:.1f}%) 扣分"

        # v6.3.20.3: sanitize strategy score (避免誤寫入欄名)
        try:
            if isinstance(safe_get(locals(), 'weight_score_raw', None, True), str):
                weight_score_raw = 0.0
        except Exception:
            pass

        # v6.3.20.4: unify squeeze pressure for TW/TWO into one field
        squeeze_pressure = ""
        try:
            squeeze_pressure = ""
        except Exception:
            squeeze_pressure = ""

        rows.append({
            "entry_date": now.strftime("%Y-%m-%d"),
            "symbol": sym,
            "market": market,
            "ticker": t,
            "name_zh": name_zh,
            "strategy": strategy,
            "strategy_desc": STRATEGY_DESC_ZH.get(strategy, strategy),
            "entry_price": round(ind["close"], 2),
            "stop_loss_price": round(ind["support_1m"], 2) if pd.notna(ind.get("support_1m")) else round(ind["close"] * 0.95, 2),
            "close": round(ind["close"], 2),
            "ma20": round(ind["ma20"], 2),
            "bias20": round(ind["bias20"], 2),
            "support_1m": round(ind["support_1m"], 2),
            "atr20": round(ind["atr20"], 4),
            "volatility_ratio": round(ind["volatility_ratio"], 4),
            "volume_ratio": round(ind["volume_ratio"], 4),
    "short_margin_ratio(%)": format_smr_display(smr, smr_meta),
    "smr_status": safe_get(locals(), "smr_status", "", SAFE_MODE),
            "otc_short_pressure": round(otc_p, 4) if market == "TWO" else None,
            "weight_mode": None,
            "weight_source_file": None,
    "weight_score_raw": (safe_float(safe_get(locals(), "weight_score_raw", 0.0, SAFE_MODE), 0.0, SAFE_MODE) + safe_float(safe_get(locals(), "score_penalty", 0.0, SAFE_MODE), 0.0, SAFE_MODE)),
            "weight_raw": None,
            "weight_prev": None,
            "weight_used": None,
            "weight_alpha": None,
            "min_trades": MIN_TRADES_FOR_WEIGHT,
            "trades_used": None,
            "annualized_pct": None,
            "mdd_pct": None,
            "position_size": None,
    "Risk Alert": (str(safe_get(locals(), "risk_alert", "", SAFE_MODE)) + (" | " + str(safe_get(locals(), "risk_note_extra", "", SAFE_MODE)) if str(safe_get(locals(), "risk_note_extra", "", SAFE_MODE)) else "")),
            "status": "OPEN",
            "exit_date": None,
            "exit_price": None,
            "pnl_pct": None,
            "hold_days": None,
        })

    if not rows:
        log('No candidates found today.')
        flush_invalid_tickers()
        return

    df = pd.DataFrame(rows)
    strategies_today = sorted(df["strategy"].unique().tolist())
    wmap, trace, mode = compute_weights_with_trace(strategies_today)

    df["weight_used"] = df["strategy"].map(wmap)
    df["strategy_weight"] = df["weight_used"]

    def tget(s, k):
        return (trace.get(s, {}) or {}).get(k)

    for col in ["weight_mode","weight_source_file","weight_score_raw","weight_raw","weight_prev","weight_alpha","trades_used","annualized_pct","mdd_pct"]:
        df[col] = df["strategy"].apply(lambda s: tget(s, col))

    df["position_size"] = df.apply(lambda x: round(calc_position_size(TOTAL_CAPITAL, float(x["weight_used"]), float(x["atr20"]), float(x["close"]), market_regime), 0), axis=1)
    df["Risk Alert"] = df["position_size"].apply(lambda p: "HIGH RISK" if p / TOTAL_CAPITAL > TH_HIGH_RISK_POS else "")

    today = _today_str()
    out_path = os.path.join(LOCAL_EXCEL_FOLDER, f"{today}_stock_selection.xlsx")
    df_view = df.sort_values(["strategy","market","symbol"]).copy()
    # === Live scoring + lights (一致) ===
    try:
        df_view = apply_live_scoring(df_view)
    except Exception as _e:
        log("apply_live_scoring skipped (error): " + repr(_e))
        # ===== Liquidity + Volatility risk block (v6.3.24) =====
    # ---- turnover_rate(%) compute (v6.3.24.3) ----
    try:
        shares_cache = os.path.join("cache", "shares_map.json")
        os.makedirs(os.path.dirname(shares_cache), exist_ok=True)
        shares_map = load_or_build_shares_map(shares_cache, max_age_hours=72)
        if shares_map:
            # volume merge from df (v6.3.24.6)
            if "volume" not in df_view.columns and "Volume" not in df_view.columns and "成交量" not in df_view.columns:
                try:
                    if "ticker" in df.columns and ("volume" in df.columns or "Volume" in df.columns or "成交量" in df.columns):
                        vsrc = "volume" if "volume" in df.columns else ("Volume" if "Volume" in df.columns else "成交量")
                        vmap = df.set_index("ticker")[vsrc].to_dict()
                        df_view["volume"] = df_view["ticker"].map(lambda x: vmap.get(x))
                except Exception as e:
                    log(f"turnover volume merge failed: {repr(e)}")
            # yahoo volume map (v6.3.27): fill candidate volume from Yahoo to guarantee turnover
            try:
                trade_date = safe_get(locals(), "selection_date", datetime.now().date(), SAFE_MODE)
                vol_cache = os.path.join("cache", f"yahoo_volume_map_{trade_date.strftime('%Y-%m-%d')}.json")
                os.makedirs(os.path.dirname(vol_cache), exist_ok=True)
                vol_map = {}
                if os.path.exists(vol_cache):
                    try:
                        with open(vol_cache, "r", encoding="utf-8") as f:
                            vol_map = json.load(f)
                    except Exception:
                        vol_map = {}
                if not vol_map:
                    vol_map = fetch_yahoo_volume_map(df_view["ticker"].tolist(), trade_date, chunk_size=80, pause_sec=1.0)
                    try:
                        with open(vol_cache, "w", encoding="utf-8") as f:
                            json.dump(vol_map, f, ensure_ascii=False)
                    except Exception:
                        pass
                log(f"yahoo volume map size (v6.3.27): {len(vol_map)}")
                if vol_map:
                    df_view["volume"] = df_view["ticker"].map(lambda t: vol_map.get(t))
            except Exception as e:
                log(f"yahoo volume map apply failed: {repr(e)}")
            df_view = compute_turnover_rate_percent(df_view, shares_map)
            log(f"turnover_rate computed: shares_map_n={len(shares_map)}")
            # turnover non-null count (v6.3.27)
            try:
                nn = pd.to_numeric(df_view.get("turnover_rate(%)"), errors="coerce").notna().sum()
                log(f"turnover non-null count (v6.3.27): {int(nn)} / {len(df_view)}")
            except Exception as e:
                log(f"turnover non-null count failed: {repr(e)}")
        else:
            df_view["turnover_rate(%)"] = float("nan")
            log("turnover_rate: shares_map empty -> NaN")
    except Exception as e:
        df_view["turnover_rate(%)"] = float("nan")
        log(f"turnover_rate compute failed: {repr(e)}")

    try:
        tcol = "turnover_rate(%)" if "turnover_rate(%)" in df_view.columns else ("周轉率(%)" if "周轉率(%)" in df_view.columns else None)
        pcol = next((c for c in ["進場價","entry_price","close","Close","收盤價","last_close"] if c in df_view.columns), None)
        aval_col = "traded_value_ntd" if "traded_value_ntd" in df_view.columns else None
        vcol_vol = "vol_annual" if "vol_annual" in df_view.columns else None
        pos_cols = [c for c in ["建議部位","position","target_position","final_position","target_weight","position_weight"] if c in df_view.columns]
        ra_col = "風險提醒" if "風險提醒" in df_view.columns else ("risk_alert" if "risk_alert" in df_view.columns else ("Risk Alert" if "Risk Alert" in df_view.columns else None))
        if tcol and pcol and pos_cols:
            pc = pos_cols[0]
            base = pd.to_numeric(df_view[pc], errors="coerce").fillna(0.0)
            tr = pd.to_numeric(df_view[tcol], errors="coerce")
            prices = pd.to_numeric(df_view[pcol], errors="coerce")
            traded_val = pd.to_numeric(df_view[aval_col], errors="coerce") if aval_col else pd.Series([float("nan")]*len(df_view))
            vol_a = pd.to_numeric(df_view[vcol_vol], errors="coerce") if vcol_vol else pd.Series([float("nan")]*len(df_view))
            mult=[]; notes=[]
            for x, pr, tv, va in zip(tr.tolist(), prices.tolist(), traded_val.tolist(), vol_a.tolist()):
                thr_turn = turnover_threshold_by_price(pr)
                thr_amt = amount_threshold_by_price(pr)
                blocked=False; note_parts=[]
                if (x is not None) and (not (isinstance(x,float) and x!=x)):
                    try:
                        if float(x) < float(thr_turn):
                            blocked=True; note_parts.append("低周轉率<" + f"{thr_turn:.2f}" + "%")
                    except Exception:
                        pass
                if (tv is not None) and (not (isinstance(tv,float) and tv!=tv)):
                    try:
                        if float(tv) < float(thr_amt):
                            blocked=True; note_parts.append("低成交值<" + f"{int(thr_amt):,}")
                    except Exception:
                        pass
                scale = vol_scale_factor(va)
                if blocked:
                    mult.append(0.0); note_parts.append("禁止下單(部位=0)")
                else:
                    mult.append(scale)
                    if scale < 1.0:
                        note_parts.append("波動控倉x" + f"{scale:.2f}")
                notes.append(" | ".join(note_parts) if note_parts else "")
            df_view[pc] = (base * pd.Series(mult)).round(6)
            if ra_col:
                add = pd.Series(notes).apply(lambda t: (" | " + t) if t else "")
                df_view[ra_col] = df_view[ra_col].astype(str) + add
    except Exception as e:
        log(f"liquidity/vol block apply failed: {repr(e)}")


    # 市場欄位中文化（僅顯示用）
    df_view["market"] = df_view["market"].map({"TW": "上市", "TWO": "上櫃"}).fillna(df_view["market"])

    # 補齊 name_zh（顯示用）
    if "name_zh" not in df_view.columns:
        df_view["name_zh"] = ""
    df_view["name_zh"] = df_view["name_zh"].fillna("")

    # fill stop_loss_price for output (v6.3.14)
    if "stop_loss_price" not in df_view.columns:
        df_view["stop_loss_price"] = pd.to_numeric(df_view.get("support_1m"), errors="coerce")
    ep = pd.to_numeric(df_view.get("entry_price"), errors="coerce")
    sl = pd.to_numeric(df_view.get("stop_loss_price"), errors="coerce")
    df_view["stop_loss_price"] = sl.fillna(pd.to_numeric(df_view.get("support_1m"), errors="coerce")).fillna((ep * 0.95).round(2))


    # ===== Full 版輸出（隱藏 Yahoo代碼欄）=====
    # ensure all columns exist for full (v6.3.13)
    for k in COLUMN_MAP_ZH.keys():
        if k not in df_view.columns:
            df_view[k] = ""


# --- compute turnover_rate(%) (v6.3.22.2) ---

    # turnover display N/A (v6.3.24.3)

    if 'turnover_rate(%)' in df_view.columns:

        trv = pd.to_numeric(df_view['turnover_rate(%)'], errors='coerce')

        df_view.loc[trv.isna() | (trv<=0), 'turnover_rate(%)'] = pd.NA  # keep numeric; display handled in df_full

    # derive squeeze_pressure from SMR (v6.3.26)


    if "short_margin_ratio" in df_view.columns and "squeeze_pressure" in df_view.columns:


        _smr = pd.to_numeric(df_view["short_margin_ratio"], errors="coerce")


        def _sp(v):


            if pd.isna(v):


                return "N/A"


            if v >= 30:


                return f"HIGH({v:.1f}%)"


            if v >= 9:


                return f"MED({v:.1f}%)"


            return ""


        df_view["squeeze_pressure"] = _smr.apply(_sp)



    df_full = df_view.reindex(columns=list(COLUMN_MAP_ZH.keys()), fill_value="").rename(columns=COLUMN_MAP_ZH)


    # turnover display N/A after df_full (v6.3.24.4)


    if "周轉率(%)" in df_full.columns:


        trv = pd.to_numeric(df_full["周轉率(%)"], errors="coerce")


        df_full["周轉率(%)"] = df_full["周轉率(%)"].astype("object")


        df_full.loc[trv.isna() | (trv<=0), "周轉率(%)"] = "N/A"
    # --- v6.3.29: 成交值/排名/周轉率顯示與部位補齊 ---
    try:
        # Turnover: keep numeric rounding (2dp) before display conversion
        if "周轉率(%)" in df_full.columns:
            _tr = pd.to_numeric(df_full["周轉率(%)"], errors="coerce")
            df_full.loc[_tr.notna(), "周轉率(%)"] = _tr.round(2)
    except Exception:
        pass

    try:
        # Trade value rank (成交值排名)
        if "成交值(元)" in df_full.columns:
            _tv = pd.to_numeric(df_full["成交值(元)"], errors="coerce")
            # If some are missing, leave blank; rank on available
            if _tv.notna().any():
                _rank = _tv.rank(ascending=False, method="min")
                # store as integer-like (no decimals)
                df_full["成交值排名"] = _rank.round(0).astype("Int64").astype("object")
                df_full.loc[_tv.isna(), "成交值排名"] = ""
    except Exception:
        pass

    # --- round numeric columns to 2dp (df_full) ---
    if "進場價" in df_full.columns:
        _x = pd.to_numeric(df_full["進場價"], errors="coerce")
        df_full["進場價"] = _x.round(2)
    if "停損價" in df_full.columns:
        _x = pd.to_numeric(df_full["停損價"], errors="coerce")
        df_full["停損價"] = _x.round(2)
    if "收盤價" in df_full.columns:
        _x = pd.to_numeric(df_full["收盤價"], errors="coerce")
        df_full["收盤價"] = _x.round(2)
    if "MA20" in df_full.columns:
        _x = pd.to_numeric(df_full["MA20"], errors="coerce")
        df_full["MA20"] = _x.round(2)
    if "乖離率(%)" in df_full.columns:
        _x = pd.to_numeric(df_full["乖離率(%)"], errors="coerce")
        df_full["乖離率(%)"] = _x.round(2)
    if "周轉率(%)" in df_full.columns:
        _x = pd.to_numeric(df_full["周轉率(%)"], errors="coerce")
        df_full["周轉率(%)"] = _x.round(2)
    if "嘎空壓力" in df_full.columns:
        _x = pd.to_numeric(df_full["嘎空壓力"], errors="coerce")
        df_full["嘎空壓力"] = _x.round(2)
    if "綜合分數" in df_full.columns:
        _x = pd.to_numeric(df_full["綜合分數"], errors="coerce")
        df_full["綜合分數"] = _x.round(2)
    if "成交值(元)" in df_full.columns:
        _x = pd.to_numeric(df_full["成交值(元)"], errors="coerce")
        df_full["成交值(元)"] = _x.round(2)

    try:
        # Position size safety: 建議部位(元) empty -> 0
        if "建議部位(元)" in df_full.columns:
            _ps = pd.to_numeric(df_full["建議部位(元)"], errors="coerce")
            df_full["建議部位(元)"] = _ps.fillna(0).round(0).astype(int)
    except Exception:
        pass

        # fixed FULL column order
    for _c in FULL_COL_ORDER_ZH:
        if _c not in df_full.columns:
            df_full[_c] = ""
    df_full = df_full.reindex(columns=FULL_COL_ORDER_ZH)

    df_full = apply_lights(df_full)

    df_full_export = apply_display_overrides(df_full)
    full_path = out_path

    df_full_export.to_excel(out_path, index=False)
    format_excel_sheet(out_path, hide_headers=["股票代號","市場"])

    # ===== 盤前決策版輸出（10欄，隱藏 Yahoo代碼欄）=====
    df_view_dec = df_view.copy()
    df_view_dec = apply_lights(df_view_dec)
    df_view_dec = apply_display_overrides(df_view_dec)
    df_view_dec = ensure_turnover_before_bias(df_view_dec)

    # Decision sheet columns (robust) (v6.3.24.2)
    decision_cols_internal = ["ticker","name_zh","strategy_desc","entry_price","stop_loss_price","turnover_rate(%)","bias20","position_size","Risk Alert"]
    
    # If df_view_dec already uses ZH columns, rename back to internal using reverse map
    rev_map = {v: k for k, v in COLUMN_MAP_ZH.items()}
    df_dec_base = df_view_dec.rename(columns=rev_map).copy()
    
    # Make sure all requested columns exist (fill empty if missing)
    for c in decision_cols_internal:
        if c not in df_dec_base.columns:
            df_dec_base[c] = ""
    
    df_decision = df_dec_base[decision_cols_internal].copy()
    df_decision = df_decision.rename(columns=COLUMN_MAP_ZH)
    
    # turnover display N/A decision (v6.3.24.4)
    if "周轉率(%)" in df_decision.columns:
        trv = pd.to_numeric(df_decision["周轉率(%)"], errors="coerce")
        df_decision["周轉率(%)"] = df_decision["周轉率(%)"].astype("object")
        df_decision.loc[trv.isna() | (trv<=0), "周轉率(%)"] = "N/A"
    base, ext = os.path.splitext(out_path)
    out_path_decision = f"{base}_決策8欄{ext}"

    # DECISION entry_date ensure
    if "進場日期" not in df_decision.columns:
        try:
            df_decision["進場日期"] = trade_date.strftime("%Y-%m-%d")
        except Exception:
            df_decision["進場日期"] = ""

    # === Decision scoring + sort + fixed columns ===
    try:
        df_decision = apply_live_scoring(df_decision)
        df_decision = apply_lights(df_decision)
        df_decision = apply_display_overrides(df_decision)
        if "綜合分數" in df_decision.columns:
            df_decision["綜合分數"] = pd.to_numeric(df_decision["綜合分數"], errors="coerce").fillna(0.0)
            _abs_bias = pd.to_numeric(df_decision.get("乖離率(%)"), errors="coerce").abs()
            df_decision = df_decision.assign(_abs_bias=_abs_bias).sort_values(["綜合分數","_abs_bias"], ascending=[False, False]).drop(columns=["_abs_bias"], errors="ignore")
    except Exception as _e:
        log("decision apply_live_scoring skipped (error): " + repr(_e))

    DECISION_COLS_FIXED = [
        "進場日期",
        "Yahoo代碼",
        "股票名稱",
        "策略說明",
        "進場價",
        "停損價",
        "乖離率(%)",
        "周轉率(%)",
        "周轉率燈號",
        "綜合分數",
        "嘎空壓力",
        "嘎空壓力燈號",
        "成交值(元)",
        "成交值排名",
        "成交值燈號",
        "建議部位(元)",
        "風險提醒",
    ]
    for _c in DECISION_COLS_FIXED:
        if _c not in df_decision.columns:
            df_decision[_c] = ""
    # --- round numeric columns to 2dp (df_decision) ---
    for _c in ["進場價","停損價","乖離率(%)","周轉率(%)","嘎空壓力","綜合分數","成交值(元)"]:
        if _c in df_decision.columns:
            _x = pd.to_numeric(df_decision[_c], errors="coerce")
            df_decision[_c] = _x.round(2)

    df_decision = df_decision.reindex(columns=DECISION_COLS_FIXED)

    df_decision.to_excel(out_path_decision, index=False)
    format_excel_sheet(out_path_decision, hide_headers=["股票代號","市場"])
    log(f"Saved decision: {out_path_decision}")
try:
    postprocess_excel(decision_path)
except Exception:
    pass

    subject = f"每日盤前檢查表 {today} ({market_regime})"
    body = f"附件為今日盤前選股結果。權重模式={mode}（已輸出 weight_* 欄位）。"
    if ENABLE_EMAIL:
        try:
            send_email_with_attachment(RECEIVER_EMAIL, out_path, subject, body)
            log('Email sent.')
        except Exception as e:
            log('Email skipped (error): ' + repr(e))
    else:
        log('Email disabled. Skipping send.')

    log(f"Saved: {out_path}")
try:
    postprocess_excel(full_path)
except Exception:
    pass
    

if __name__ == "__main__":
    import traceback
    from pathlib import Path
    from datetime import datetime
    import sys

    Path("logs").mkdir(exist_ok=True)
    try:
        main()
    except SystemExit as e:
        code = getattr(e, "code", 1)
        if code not in (0, None):
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            p = Path("logs") / f"daily_auto_run_error_{ts}.log"
            content = "SystemExit: " + repr(code) + "\n\n" + "STACK:\n" + "".join(traceback.format_stack())
            p.write_text(content, encoding="utf-8")
            print(f"[FATAL] saved traceback: {p}")
        raise
    except BaseException:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        p = Path("logs") / f"daily_auto_run_error_{ts}.log"
        p.write_text(traceback.format_exc(), encoding="utf-8")
        print(f"[FATAL] saved traceback: {p}")
        raise
