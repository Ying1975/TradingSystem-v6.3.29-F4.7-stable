
# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import glob
from pathlib import Path

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from lights_unified import apply_lights, apply_display_overrides

TOP20_COL_ORDER = [
    "進場日期",
    "Yahoo代碼",
    "股票名稱",
    "策略說明",
    "進場價",
    "停損價",
    "嘎空壓力",
    "嘎空壓力燈號",
    "乖離率(%)",
    "周轉率(%)",
    "周轉率燈號",
    "成交值(元)",
    "成交值排名",
    "成交值燈號",
    "建議部位(元)",
    "風險提醒",
    "綜合分數",
]

def pick_latest_full(records_dir: str, run_date: str | None):
    if run_date:
        cand = os.path.join(records_dir, f"{run_date}_stock_selection.xlsx")
        if os.path.exists(cand):
            return cand
    files = sorted(glob.glob(os.path.join(records_dir, "*_stock_selection.xlsx")))
    return files[-1] if files else None

def apply_alignment_and_lights(wb):
    color_map = {"🔴": "FFFF0000", "🟡": "FFFFA500", "🟢": "FF00AA00"}
    fill_map  = {"🔴": "FFFFE5E5", "🟡": "FFFFF2CC", "🟢": "FFE2F0D9"}
    na_fill = "FFF2F2F2"

    align = Alignment(horizontal="right", vertical="bottom")
    emoji_align = Alignment(horizontal="center", vertical="bottom")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align

        headers = {str(c.value): c.column for c in ws[1] if c.value is not None}
        for h, col in headers.items():
            if "燈號" not in h:
                continue
            for r in range(2, ws.max_row + 1):
                c = ws.cell(row=r, column=col)
                v = c.value
                if v in color_map:
                    c.font = Font(color=color_map[v], bold=True)
                    c.fill = PatternFill("solid", fgColor=fill_map.get(v, na_fill))
                    c.alignment = emoji_align
                elif v in ("N/A", "LOW"):
                    c.font = Font(color="FF666666", bold=True)
                    c.fill = PatternFill("solid", fgColor=na_fill)
                    c.alignment = emoji_align

def postprocess_excel(path: str):
    try:
        wb = load_workbook(path)
        apply_alignment_and_lights(wb)
        wb.save(path)
    except Exception:
        return

def build_top20(df_full: pd.DataFrame) -> pd.DataFrame:
    if df_full is None or len(df_full)==0:
        return pd.DataFrame(columns=TOP20_COL_ORDER)

    df = df_full.copy()

    mapping = {
        "entry_date": "進場日期",
        "ticker": "Yahoo代碼",
        "name_zh": "股票名稱",
        "strategy_desc": "策略說明",
        "entry_price": "進場價",
        "stop_loss_price": "停損價",
        "bias20": "乖離率(%)",
        "turnover_rate(%)": "周轉率(%)",
        "turnover_rate": "周轉率(%)",
        "trade_value": "成交值(元)",
        "成交值(元)": "成交值(元)",
        "trade_value_rank": "成交值排名",
        "成交值排名": "成交值排名",
        "position_size": "建議部位(元)",
        "Risk Alert": "風險提醒",
        "risk_alert": "風險提醒",
        "final_score": "綜合分數",
        "squeeze_pressure": "嘎空壓力",
        "嘎空壓力": "嘎空壓力",
    }
    for src, dst in mapping.items():
        if src in df.columns and dst not in df.columns:
            df[dst] = df[src]

    for c in TOP20_COL_ORDER:
        if c not in df.columns:
            df[c] = ""

    df = apply_lights(df)
    df = apply_display_overrides(df)

    score = pd.to_numeric(df.get("綜合分數"), errors="coerce")
    bias = pd.to_numeric(df.get("乖離率(%)"), errors="coerce")
    df["__score"] = score
    df["__bias"] = bias
    df = df.sort_values(["__score","__bias"], ascending=[False, True], na_position="last")

    top = df.head(20).copy()
    top = top.reindex(columns=TOP20_COL_ORDER)
    return top

def main():
    base_dir = os.path.dirname(__file__)
    records_dir = os.path.join(base_dir, "daily_excel_records")
    os.makedirs(records_dir, exist_ok=True)

    run_date = os.environ.get("RUN_DATE","").strip() or None
    full_path = pick_latest_full(records_dir, run_date)

    out_path = os.environ.get("TOP20_OUT_PATH","").strip()
    if not out_path:
        date_tag = run_date or (os.path.basename(full_path).split("_stock_selection.xlsx")[0] if full_path else "UNKNOWN")
        out_path = os.path.join(records_dir, f"{date_tag}_Top20_推薦清單.xlsx")

    if not full_path:
        pd.DataFrame(columns=TOP20_COL_ORDER).to_excel(out_path, index=False)
        postprocess_excel(out_path)
        return

    df_full = pd.read_excel(full_path)
    top20 = build_top20(df_full)

    top20.to_excel(out_path, index=False)
    postprocess_excel(out_path)

if __name__ == "__main__":
    main()
